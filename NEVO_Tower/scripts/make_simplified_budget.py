"""
Create Simplified NEVO Tower Budget - Consolidate 23 CSI divisions into 7 major categories
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
import json

# Load base data
with open('data/nevo_parsed.json', 'r') as f:
    project_data = json.load(f)

TOTAL_GSF = project_data['building_totals']['total_gsf']

def create_simplified_budget():
    """Create simplified budget with consolidated categories"""
    
    wb = openpyxl.load_workbook('outputs/NEVO_Interactive_Budget_V3.xlsx')
    
    # Remove old Hard Costs sheet and create new simplified one
    if 'Hard Costs - Simplified' in wb.sheetnames:
        del wb['Hard Costs - Simplified']
    
    ws = wb.create_sheet("Hard Costs - Simplified", 2)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    editable_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    ws['A1'] = 'HARD COSTS - SIMPLIFIED (7 MAJOR CATEGORIES)'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    
    ws['A2'] = 'Consolidated from 23 CSI divisions for clarity'
    ws.merge_cells('A2:E2')
    
    row = 4
    headers = ['Category', 'Description', 'Rate ($/GSF)', 'GSF', 'Total Cost']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    row += 1
    
    # Simplified categories with combined rates
    categories = [
        ('1', 'SITE & FOUNDATION', 
         'Site work, demolition, earthwork, utilities connections (CSI 01, 02, 31)',
         18.89),  # 10.53 + 4.56 + 3.68 + utilities
        
        ('2', 'STRUCTURE', 
         'Concrete structure, structural steel, framing (CSI 03, 05)',
         46.92),  # 39.47 + 7.45
        
        ('3', 'BUILDING ENVELOPE', 
         'Masonry, waterproofing, roofing, windows, doors (CSI 04, 07, 08)',
         46.31),  # 2.46 + 19.30 + 24.55
        
        ('4', 'INTERIOR FINISHES', 
         'Drywall, flooring, paint, millwork, cabinets, specialties (CSI 06, 09, 10, 11, 12, 13)',
         48.56),  # 2.81 + 36.83 + 3.95 + 3.33 + 1.05 + 1.58 - overlaps
        
        ('5', 'MECHANICAL, PLUMBING & FIRE PROTECTION', 
         'HVAC, plumbing, fire sprinklers (CSI 21, 22, 23)',
         28.50),  # 6.58 + 8.77 + 13.15
        
        ('6', 'ELECTRICAL & LOW VOLTAGE', 
         'Power, lighting, data, phone, security, fire alarm (CSI 26, 27, 28)',
         16.65),  # 13.15 + 1.75 + 1.75
        
        ('7', 'ELEVATORS & SITE IMPROVEMENTS', 
         'Elevators, landscaping, paving, site utilities (CSI 14, 32, 33)',
         9.13),  # 2.72 + 3.51 + 2.81 + 14 (elevators adjusted)
    ]
    
    start_row = row
    for num, category, description, rate in categories:
        ws[f'A{row}'] = num
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = category
        ws[f'B{row}'].font = Font(bold=True, size=11)
        ws[f'C{row}'] = rate
        ws[f'C{row}'].fill = editable_fill
        ws[f'C{row}'].number_format = '$#,##0.00'
        ws[f'D{row}'] = TOTAL_GSF
        ws[f'D{row}'].number_format = '#,##0'
        ws[f'E{row}'] = f'=C{row}*D{row}'
        ws[f'E{row}'].number_format = '$#,##0'
        ws[f'E{row}'].font = Font(bold=True)
        row += 1
        
        # Add description row
        ws[f'B{row}'] = description
        ws[f'B{row}'].font = Font(italic=True, size=9)
        ws.merge_cells(f'B{row}:E{row}')
        row += 1
    
    # Subtotal
    row += 1
    ws[f'B{row}'] = 'SUBTOTAL - Direct Construction Costs'
    ws[f'B{row}'].font = Font(bold=True, size=12)
    ws[f'E{row}'] = f'=SUM(E{start_row},E{start_row+2},E{start_row+4},E{start_row+6},E{start_row+8},E{start_row+10},E{start_row+12})'
    ws[f'E{row}'].number_format = '$#,##0'
    ws[f'E{row}'].fill = total_fill
    ws[f'E{row}'].font = Font(bold=True, size=12)
    subtotal_row = row
    row += 2
    
    # GC COSTS
    ws[f'B{row}'] = 'GENERAL CONDITIONS (9%)'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = '=GC_GC'
    ws[f'C{row}'].number_format = '0.0%'
    ws[f'E{row}'] = f'=E{subtotal_row}*GC_GC'
    ws[f'E{row}'].number_format = '$#,##0'
    gc_row = row
    row += 1
    
    ws[f'B{row}'] = '  Site office, supervision, permits, safety'
    ws[f'B{row}'].font = Font(italic=True, size=9)
    ws.merge_cells(f'B{row}:E{row}')
    row += 1
    
    ws[f'B{row}'] = 'OVERHEAD & PROFIT (10%)'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = '=GC_OHP'
    ws[f'C{row}'].number_format = '0.0%'
    ws[f'E{row}'] = f'=(E{subtotal_row}+E{gc_row})*GC_OHP'
    ws[f'E{row}'].number_format = '$#,##0'
    ohp_row = row
    row += 1
    
    ws[f'B{row}'] = "  GC's business costs, insurance, profit margin"
    ws[f'B{row}'].font = Font(italic=True, size=9)
    ws.merge_cells(f'B{row}:E{row}')
    row += 2
    
    # Before contingency
    ws[f'B{row}'] = 'SUBTOTAL before Contingencies'
    ws[f'B{row}'].font = Font(bold=True, size=12)
    ws[f'E{row}'] = f'=E{subtotal_row}+E{gc_row}+E{ohp_row}'
    ws[f'E{row}'].number_format = '$#,##0'
    ws[f'E{row}'].fill = total_fill
    ws[f'E{row}'].font = Font(bold=True, size=12)
    before_cont_row = row
    row += 2
    
    # CONTINGENCIES
    ws[f'B{row}'] = 'DESIGN CONTINGENCY (8%)'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = '=Design_Cont'
    ws[f'C{row}'].number_format = '0.0%'
    ws[f'E{row}'] = f'=E{before_cont_row}*Design_Cont'
    ws[f'E{row}'].number_format = '$#,##0'
    design_cont_row = row
    row += 1
    
    ws[f'B{row}'] = '  Buffer for design changes during construction'
    ws[f'B{row}'].font = Font(italic=True, size=9)
    ws.merge_cells(f'B{row}:E{row}')
    row += 1
    
    ws[f'B{row}'] = 'CONSTRUCTION CONTINGENCY (5%)'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = '=Constr_Cont'
    ws[f'C{row}'].number_format = '0.0%'
    ws[f'E{row}'] = f'=E{before_cont_row}*Constr_Cont'
    ws[f'E{row}'].number_format = '$#,##0'
    constr_cont_row = row
    row += 1
    
    ws[f'B{row}'] = '  Unexpected field conditions (rock, utilities, etc.)'
    ws[f'B{row}'].font = Font(italic=True, size=9)
    ws.merge_cells(f'B{row}:E{row}')
    row += 1
    
    ws[f'B{row}'] = 'OWNER CONTINGENCY (3%)'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = '=Owner_Cont'
    ws[f'C{row}'].number_format = '0.0%'
    ws[f'E{row}'] = f'=E{before_cont_row}*Owner_Cont'
    ws[f'E{row}'].number_format = '$#,##0'
    owner_cont_row = row
    row += 1
    
    ws[f'B{row}'] = "  Owner's reserve for changes/upgrades"
    ws[f'B{row}'].font = Font(italic=True, size=9)
    ws.merge_cells(f'B{row}:E{row}')
    row += 2
    
    # TOTAL HARD COSTS
    ws[f'B{row}'] = 'TOTAL HARD COSTS'
    ws[f'B{row}'].font = Font(bold=True, size=14)
    ws[f'E{row}'] = f'=E{before_cont_row}+E{design_cont_row}+E{constr_cont_row}+E{owner_cont_row}'
    ws[f'E{row}'].number_format = '$#,##0'
    ws[f'E{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    ws[f'E{row}'].font = Font(bold=True, size=14)
    total_row = row
    row += 1
    
    # Cost per GSF
    ws[f'B{row}'] = 'Cost per GSF'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'E{row}'] = f'=E{total_row}/{TOTAL_GSF}'
    ws[f'E{row}'].number_format = '$#,##0.00'
    ws[f'E{row}'].font = Font(bold=True)
    row += 2
    
    # Comparison
    ws[f'B{row}'] = 'Detailed CSI Budget (23 divisions)'
    ws[f'E{row}'] = '=TotalHardCosts'
    ws[f'E{row}'].number_format = '$#,##0'
    row += 1
    
    ws[f'B{row}'] = 'Simplified Budget (7 categories)'
    ws[f'E{row}'] = f'=E{total_row}'
    ws[f'E{row}'].number_format = '$#,##0'
    row += 1
    
    ws[f'B{row}'] = 'Difference'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'E{row}'] = f'=E{row-2}-E{row-1}'
    ws[f'E{row}'].number_format = '$#,##0'
    ws[f'E{row}'].font = Font(bold=True)
    
    # Save
    filename = 'outputs/NEVO_Simplified_Budget.xlsx'
    wb.save(filename)
    
    print(f"\n✅ Simplified Budget created: {filename}")
    print("\n7 Major Categories (reduced from 23 CSI divisions):")
    print("1. Site & Foundation")
    print("2. Structure")
    print("3. Building Envelope")
    print("4. Interior Finishes")
    print("5. Mechanical, Plumbing & Fire Protection")
    print("6. Electrical & Low Voltage")
    print("7. Elevators & Site Improvements")
    
    return filename

if __name__ == "__main__":
    create_simplified_budget()
