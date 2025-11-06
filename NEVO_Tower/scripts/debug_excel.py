import openpyxl

wb = openpyxl.load_workbook('outputs/NEVO_Interactive_Budget.xlsx')
ws = wb['Hard Costs']

print("Hard Costs sheet formulas:")
print(f"E29 (GC): {ws['E29'].value}")
print(f"E30 (O&P): {ws['E30'].value}")
print(f"E32 (before cont): {ws['E32'].value}")

print("\nAssumptions sheet layout:")
ws2 = wb['Assumptions']
for i in range(10, 26):
    print(f"Row {i}: {ws2[f'A{i}'].value} = {ws2[f'B{i}'].value}")
