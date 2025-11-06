#!/usr/bin/env python3
"""
Add sheets from _NEVO TOWER 50 UNITS (1).xlsx to NEVO_Interactive_Budget_V3.xlsx
"""

import openpyxl
from copy import copy

def copy_sheet(source_sheet, target_workbook):
    """Copy a sheet from source to target workbook, preserving formatting"""
    # Create new sheet with the same name
    target_sheet = target_workbook.create_sheet(source_sheet.title)
    
    # Copy column dimensions
    for col_letter, col_dim in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[col_letter].width = col_dim.width
    
    # Copy row dimensions
    for row_num, row_dim in source_sheet.row_dimensions.items():
        target_sheet.row_dimensions[row_num].height = row_dim.height
    
    # Copy all cells with formatting
    for row in source_sheet.iter_rows():
        for cell in row:
            target_cell = target_sheet[cell.coordinate]
            
            # Copy value
            if cell.value:
                target_cell.value = cell.value
            
            # Copy formatting
            if cell.has_style:
                target_cell.font = copy(cell.font)
                target_cell.border = copy(cell.border)
                target_cell.fill = copy(cell.fill)
                target_cell.number_format = copy(cell.number_format)
                target_cell.protection = copy(cell.protection)
                target_cell.alignment = copy(cell.alignment)
    
    # Copy merged cells
    for merged_cell_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_cell_range))
    
    return target_sheet

def main():
    # Load source workbook
    source_path = r'C:\Users\17274\Downloads\_NEVO TOWER 50 UNITS (1).xlsx'
    source_wb = openpyxl.load_workbook(source_path)
    
    # Load target workbook
    target_path = r'C:\Users\17274\ME\2829-Niagara-Street\NEVO_Tower\outputs\NEVO_Interactive_Budget_V3.xlsx'
    target_wb = openpyxl.load_workbook(target_path)
    
    print(f"Source sheets: {source_wb.sheetnames}")
    print(f"Target sheets before: {target_wb.sheetnames}")
    
    # Copy all sheets from source
    for sheet_name in source_wb.sheetnames:
        print(f"\nCopying sheet: {sheet_name}")
        
        # Remove if sheet already exists in target
        if sheet_name in target_wb.sheetnames:
            print(f"  Removing existing sheet: {sheet_name}")
            del target_wb[sheet_name]
        
        # Copy the sheet
        source_sheet = source_wb[sheet_name]
        copy_sheet(source_sheet, target_wb)
        print(f"  ✓ Copied {sheet_name}")
    
    # Save target workbook
    target_wb.save(target_path)
    print(f"\n✓ Saved updated workbook to: {target_path}")
    print(f"Final sheets: {target_wb.sheetnames}")

if __name__ == '__main__':
    main()
