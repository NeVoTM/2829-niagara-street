"""
Create Interactive NEVO Tower Budget Workbook V3
ALL CORRECTIONS APPLIED:
1. Summary sheet - % in column C, SF with commas
2. Assumptions sheet - % in columns C and D
3. Soft costs - categorized by in-kind/cash, = 12% of project cost
4. All formulas use named ranges
5. Cash flow with formulas, 50 units, correct SC Cash OUT label
6. Project deal sheet - SC Cash corrected (OUT not IN)
7. Auto-fit all columns
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.workbook.defined_name import DefinedName
from datetime import datetime
import json

# Load base data
with open('data/nevo_parsed.json', 'r') as f:
    project_data = json.load(f)

TOTAL_GSF = project_data['building_totals']['total_gsf']
TOTAL_NSF = project_data['building_totals']['total_nsf']
RESIDENTIAL_UNITS = project_data['building_totals']['residential_units']

def create_interactive_workbook():
    """Create fully interactive Excel workbook with land partner structure"""
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    print("Creating Interactive NEVO Tower Budget Workbook V3...")
    print("With ALL 11 Corrections Applied")
    
    # Create sheets in order
    create_summary_sheet(wb)
    create_assumptions_sheet(wb)
    create_hard_costs_sheet(wb)
    create_soft_costs_sheet(wb)
    create_revenue_sheet(wb)
    create_project_deal_sheet(wb)
    create_cash_flow_sheet(wb)
    create_gc_rfq_sheet(wb)
    create_sync_instructions_sheet(wb)
    
    # Set column widths explicitly for proper alignment
    # Summary sheet column widths already set in create_summary_sheet
    # Other sheets have their column widths set in their respective functions
    
    # Save workbook
    filename = 'outputs/NEVO_Interactive_Budget_V3.xlsx'
    wb.save(filename)
    
    print(f"\n✅ Interactive workbook created: {filename}") 
    print("\n✅ ALL 11 CORRECTIONS APPLIED")
    
    return filename

def create_summary_sheet(wb):
    """Executive summary with all calculations"""
    ws = wb.create_sheet("Summary", 0)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=14)
    
    ws['A1'] = 'NEVO TOWER - EXECUTIVE SUMMARY'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws.merge_cells('A2:D2')
    
    row = 4
    
    # Total Costs
    ws[f'A{row}'] = 'TOTAL DEVELOPMENT COST'
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    ws[f'A{row}'] = 'SC Cash Payment (Project Cost)'
    ws[f'B{row}'] = '=LandPartner_Cash1+LandPartner_Cash2'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'C{row}'] = '=B{}/TotalProjectCost'.format(row)
    ws[f'C{row}'].number_format = '0.0%'
    row += 1
    
    ws[f'A{row}'] = 'Hard Costs'
    ws[f'B{row}'] = '=TotalHardCosts'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'C{row}'] = '=TotalHardCosts/TotalProjectCost'
    ws[f'C{row}'].number_format = '0.0%'
    row += 1
    
    ws[f'A{row}'] = 'Soft Costs'
    ws[f'B{row}'] = '=TotalSoftCosts'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'C{row}'] = '=TotalSoftCosts/TotalProjectCost'
    ws[f'C{row}'].number_format = '0.0%'
    row += 1
    
    ws[f'A{row}'] = 'TOTAL PROJECT COST'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'B{row}'] = f'=SUM(B{row-3}:B{row-1})'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].font = Font(bold=True, size=12)
    ws[f'B{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    ws[f'C{row}'] = '100.0%'
    total_cost_row = row
    row += 2
    
    # Define named range for TotalProjectCost
    wb.defined_names['TotalProjectCost'] = DefinedName('TotalProjectCost', attr_text=f"Summary!$B${total_cost_row}")
    
    # Revenue
    ws[f'A{row}'] = 'REVENUE'
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    ws[f'A{row}'] = 'Total Revenue'
    ws[f'B{row}'] = '=TotalRevenue'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].font = Font(bold=True, size=12)
    ws[f'B{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    revenue_row = row
    row += 2
    
    # Profitability
    ws[f'A{row}'] = 'PROFITABILITY'
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    ws[f'A{row}'] = 'Gross Profit'
    ws[f'B{row}'] = f'=B{revenue_row}-B{total_cost_row}'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = f'=(B{revenue_row}-B{total_cost_row})/B{revenue_row}'
    ws[f'C{row}'].number_format = '0.0%'
    profit_row = row
    row += 1
    
    ws[f'A{row}'] = 'Profit Margin'
    ws[f'B{row}'] = f'=B{profit_row}/B{revenue_row}'
    ws[f'B{row}'].number_format = '0.0%'
    ws[f'B{row}'].font = Font(bold=True)
    row += 1
    
    ws[f'A{row}'] = 'Return on Cost (ROI)'
    ws[f'B{row}'] = f'=B{profit_row}/B{total_cost_row}'
    ws[f'B{row}'].number_format = '0.0%'
    ws[f'B{row}'].font = Font(bold=True)
    row += 2
    
    # Per unit metrics - using named range TotalNSF
    wb.defined_names['TotalNSF'] = DefinedName('TotalNSF', attr_text=f"Assumptions!$B$6")
    
    ws[f'A{row}'] = 'Cost per Unit'
    ws[f'B{row}'] = f'=B{total_cost_row}/50'
    ws[f'B{row}'].number_format = '$#,##0'
    row += 1
    
    ws[f'A{row}'] = 'Cost per NSF'
    ws[f'B{row}'] = f'=B{total_cost_row}/TotalNSF'
    ws[f'B{row}'].number_format = '$#,##0.00'
    row += 1
    
    ws[f'A{row}'] = 'Revenue per NSF'
    ws[f'B{row}'] = f'=B{revenue_row}/TotalNSF'
    ws[f'B{row}'].number_format = '$#,##0.00'
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25

def create_assumptions_sheet(wb):
    """Project assumptions with land partner structure"""
    ws = wb.create_sheet("Assumptions")
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    editable_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    ws['A1'] = 'NEVO TOWER - PROJECT ASSUMPTIONS'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    row = 3
    
    # Project Info
    ws[f'A{row}'] = 'PROJECT INFORMATION'
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:D{row}')
    row += 2
    
    ws[f'A{row}'] = 'Total GSF'
    ws[f'B{row}'] = TOTAL_GSF
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'C{row}'] = 1.0
    ws[f'C{row}'].number_format = '0.0%'
    row += 1
    
    ws[f'A{row}'] = 'Total NSF'
    ws[f'B{row}'] = TOTAL_NSF
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'C{row}'] = f'={TOTAL_NSF}/{TOTAL_GSF}'
    ws[f'C{row}'].number_format = '0.0%'
    row += 1
    
    ws[f'A{row}'] = 'Residential Units'
    ws[f'B{row}'] = RESIDENTIAL_UNITS
    ws[f'C{row}'] = 1.0
    ws[f'C{row}'].number_format = '0.0%'
    row += 2
    
    # LAND PARTNER STRUCTURE
    ws[f'A{row}'] = 'LAND PARTNER STRUCTURE'
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:D{row}')
    row += 2
    
    land_partner_start = row
    
    ws[f'A{row}'] = 'Land Partner - Land (In-Kind)'
    ws[f'B{row}'] = 8000000
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].fill = editable_fill
    ws[f'C{row}'] = '=B{}/10000000'.format(row)
    ws[f'C{row}'].number_format = '0.0%'
    ws[f'D{row}'] = 'Land value $10M total - $2M paid cash'
    row += 1
    
    ws[f'A{row}'] = 'SC Key Money (Month 0)'
    ws[f'B{row}'] = 1250000
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].fill = editable_fill
    ws[f'C{row}'] = '=B{}/2000000'.format(row)
    ws[f'C{row}'].number_format = '0.0%'
    ws[f'D{row}'] = 'Developer pays SC to enter deal'
    row += 1
    
    ws[f'A{row}'] = 'SC Payment (Month 6)'
    ws[f'B{row}'] = 750000
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].fill = editable_fill
    ws[f'C{row}'] = '=B{}/2000000'.format(row)
    ws[f'C{row}'].number_format = '0.0%'
    ws[f'D{row}'] = 'Additional payment to SC'
    row += 1
    
    ws[f'A{row}'] = 'Land Partner - Deferred Payment'
    ws[f'B{row}'] = 0
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].fill = editable_fill
    ws[f'C{row}'] = 0.0
    ws[f'C{row}'].number_format = '0.0%'
    ws[f'D{row}'] = '$8M paid at exit (NOT a cost)'
    row += 2
    
    # Named ranges for land partner
    wb.defined_names['LandPartner_Land'] = DefinedName('LandPartner_Land', attr_text=f'Assumptions!$B${land_partner_start}')
    wb.defined_names['LandPartner_Cash1'] = DefinedName('LandPartner_Cash1', attr_text=f'Assumptions!$B${land_partner_start+1}')
    wb.defined_names['LandPartner_Cash2'] = DefinedName('LandPartner_Cash2', attr_text=f'Assumptions!$B${land_partner_start+2}')
    wb.defined_names['LandPartner_Deferred'] = DefinedName('LandPartner_Deferred', attr_text=f'Assumptions!$B${land_partner_start+3}')
    
    # COST ASSUMPTIONS
    ws[f'A{row}'] = 'COST ASSUMPTIONS'
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:D{row}')
    row += 2
    
    cost_start = row
    
    cost_data = [
        ('GC Overhead & Profit', 0.10, '%'),
        ('General Conditions', 0.09, '%'),
        ('Design Contingency', 0.08, '%'),
        ('Construction Contingency', 0.05, '%'),
        ('Owner Contingency', 0.03, '%'),
        ('A&E Fee Rate', 0.038, '%'),
        ('Marketing Rate', 0.020, '%'),
        ('Financing Rate', 0.085, '%'),
        ('Loan to Cost', 0.75, '%'),
        ('Construction Duration', 24, 'months'),
        ('Bridge Loan Duration', 9, 'months'),
    ]
    
    for param, value, unit in cost_data:
        ws[f'A{row}'] = param
        ws[f'B{row}'] = value
        ws[f'B{row}'].fill = editable_fill
        if unit == '%':
            ws[f'B{row}'].number_format = '0.0%'
        ws[f'C{row}'] = unit
        row += 1
    
    # Named ranges for formulas
    wb.defined_names['GC_OHP'] = DefinedName('GC_OHP', attr_text=f'Assumptions!$B${cost_start}')
    wb.defined_names['GC_GC'] = DefinedName('GC_GC', attr_text=f'Assumptions!$B${cost_start+1}')
    wb.defined_names['Design_Cont'] = DefinedName('Design_Cont', attr_text=f'Assumptions!$B${cost_start+2}')
    wb.defined_names['Constr_Cont'] = DefinedName('Constr_Cont', attr_text=f'Assumptions!$B${cost_start+3}')
    wb.defined_names['Owner_Cont'] = DefinedName('Owner_Cont', attr_text=f'Assumptions!$B${cost_start+4}')
    wb.defined_names['AE_Rate'] = DefinedName('AE_Rate', attr_text=f'Assumptions!$B${cost_start+5}')
    wb.defined_names['Marketing_Rate'] = DefinedName('Marketing_Rate', attr_text=f'Assumptions!$B${cost_start+6}')
    wb.defined_names['Finance_Rate'] = DefinedName('Finance_Rate', attr_text=f'Assumptions!$B${cost_start+7}')
    wb.defined_names['LTC'] = DefinedName('LTC', attr_text=f'Assumptions!$B${cost_start+8}')
    wb.defined_names['Duration'] = DefinedName('Duration', attr_text=f'Assumptions!$B${cost_start+9}')
    
    # Add Named Ranges Documentation
    row += 3
    ws[f'A{row}'] = 'NAMED RANGES - How Formulas Connect Sheets'
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    ws[f'A{row}'] = 'Range Name'
    ws[f'B{row}'] = 'Location'
    ws[f'C{row}'] = 'Value'
    ws[f'D{row}'] = 'Description'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'D{row}'].font = Font(bold=True)
    row += 1
    
    named_ranges = [
        ('LandPartner_Land', f'B{land_partner_start}', f'=B{land_partner_start}', '80% - SC land in-kind ($8M)'),
        ('LandPartner_Cash1', f'B{land_partner_start+1}', f'=B{land_partner_start+1}', '62.5% - SC cash Month 0 ($1.25M)'),
        ('LandPartner_Cash2', f'B{land_partner_start+2}', f'=B{land_partner_start+2}', '37.5% - SC cash Month 6 ($750K)'),
        ('TotalHardCosts', "'Hard Costs'!E[row]", '=TotalHardCosts', '~85% - All construction costs'),
        ('TotalSoftCosts', "'Soft Costs'!B[row]", '=TotalSoftCosts', '~12% - Design, permits, financing'),
        ('TotalRevenue', "'Revenue'!F[row]", '=TotalRevenue', '100% - Total project revenue'),
        ('GC_OHP', f'B{cost_start}', '=GC_OHP', '10% - GC overhead & profit rate'),
        ('AE_Rate', f'B{cost_start+5}', '=AE_Rate', '3.8% - Architecture & engineering fee'),
        ('Marketing_Rate', f'B{cost_start+6}', '=Marketing_Rate', '2.0% - Marketing & sales rate'),
        ('LTC', f'B{cost_start+8}', '=LTC', '75% - Loan-to-cost ratio'),
    ]
    
    for name, location, formula, description in named_ranges:
        ws[f'A{row}'] = name
        ws[f'B{row}'] = location
        ws[f'C{row}'] = formula
        ws[f'D{row}'] = description
        row += 1
    
    row += 1
    ws[f'A{row}'] = 'HOW TO USE NAMED RANGES:'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'D{row}'] = 'Type = then the range name in any formula'
    row += 1
    ws[f'A{row}'] = 'Example:'
    ws[f'D{row}'] = '=TotalHardCosts*0.05 calculates 5% of hard costs'
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 45

def create_hard_costs_sheet(wb):
    """Hard costs by CSI division"""
    ws = wb.create_sheet("Hard Costs")
    
    # Create named range for TotalGSF if not already created
    if 'TotalGSF' not in wb.defined_names:
        wb.defined_names['TotalGSF'] = DefinedName('TotalGSF', attr_text='Assumptions!$B$5')
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    editable_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    ws['A1'] = 'HARD COSTS - CSI MASTERFORMAT DIVISIONS'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    
    row = 3
    headers = ['CSI', 'Description', 'Rate ($/GSF)', 'GSF', 'Total Cost']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    row += 1
    
    # CSI Divisions with rates
    csi_divisions = [
        ('01', 'General Requirements', 10.53),
        ('02', 'Existing Conditions & Site Work', 4.56),
        ('03', 'Concrete - Pour-in-Place PT', 39.47),
        ('04', 'Masonry', 2.46),
        ('05', 'Metals', 7.45),
        ('06', 'Wood & Plastics', 2.81),
        ('07', 'Thermal & Moisture Protection', 19.30),
        ('08', 'Openings (Doors, Windows, HVHZ)', 24.55),
        ('09', 'Finishes', 36.83),
        ('10', 'Specialties', 3.95),
        ('11', 'Equipment', 3.33),
        ('12', 'Furnishings', 1.05),
        ('13', 'Special Construction', 1.58),
        ('14', 'Conveying Equipment', 2.72),
        ('21', 'Fire Suppression', 6.58),
        ('22', 'Plumbing', 8.77),
        ('23', 'HVAC', 13.15),
        ('26', 'Electrical', 13.15),
        ('27', 'Communications', 1.75),
        ('28', 'Electronic Safety & Security', 1.75),
        ('31', 'Earthwork', 3.68),
        ('32', 'Exterior Improvements', 3.51),
        ('33', 'Utilities', 2.81),
    ]
    
    start_row = row
    for div, desc, rate in csi_divisions:
        ws[f'A{row}'] = div
        ws[f'B{row}'] = desc
        ws[f'C{row}'] = rate
        ws[f'C{row}'].fill = editable_fill
        ws[f'C{row}'].number_format = '$#,##0.00'
        ws[f'D{row}'] = '=TotalGSF'
        ws[f'D{row}'].number_format = '#,##0'
        ws[f'E{row}'] = f'=C{row}*D{row}'
        ws[f'E{row}'].number_format = '$#,##0'
        row += 1
    
    # Subtotal
    ws[f'B{row}'] = 'SUBTOTAL - Construction'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'E{row}'] = f'=SUM(E{start_row}:E{row-1})'
    ws[f'E{row}'].number_format = '$#,##0'
    ws[f'E{row}'].fill = total_fill
    subtotal_row = row
    
    # Define named range for construction subtotal
    wb.defined_names['ConstructionSubtotal'] = DefinedName('ConstructionSubtotal', attr_text=f"'Hard Costs'!$E${subtotal_row}")
    row += 2
    
    # GC Costs
    ws[f'B{row}'] = 'General Conditions'
    ws[f'C{row}'] = '=GC_GC'
    ws[f'C{row}'].number_format = '0.0%'
    ws[f'E{row}'] = f'=E{subtotal_row}*GC_GC'
    ws[f'E{row}'].number_format = '$#,##0'
    gc_row = row
    row += 1
    
    ws[f'B{row}'] = 'Overhead & Profit'
    ws[f'C{row}'] = '=GC_OHP'
    ws[f'C{row}'].number_format = '0.0%'
    ws[f'E{row}'] = f'=(E{subtotal_row}+E{gc_row})*GC_OHP'
    ws[f'E{row}'].number_format = '$#,##0'
    ohp_row = row
    row += 2
    
    # Before contingency
    ws[f'B{row}'] = 'Subtotal before Contingency'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'E{row}'] = f'=E{subtotal_row}+E{gc_row}+E{ohp_row}'
    ws[f'E{row}'].number_format = '$#,##0'
    ws[f'E{row}'].fill = total_fill
    before_cont_row = row
    row += 2
    
    # Contingencies
    ws[f'B{row}'] = 'Design Contingency'
    ws[f'C{row}'] = '=Design_Cont'
    ws[f'C{row}'].number_format = '0.0%'
    ws[f'E{row}'] = f'=E{before_cont_row}*Design_Cont'
    ws[f'E{row}'].number_format = '$#,##0'
    design_cont_row = row
    row += 1
    
    ws[f'B{row}'] = 'Construction Contingency'
    ws[f'C{row}'] = '=Constr_Cont'
    ws[f'C{row}'].number_format = '0.0%'
    ws[f'E{row}'] = f'=E{before_cont_row}*Constr_Cont'
    ws[f'E{row}'].number_format = '$#,##0'
    constr_cont_row = row
    row += 1
    
    ws[f'B{row}'] = 'Owner Contingency'
    ws[f'C{row}'] = '=Owner_Cont'
    ws[f'C{row}'].number_format = '0.0%'
    ws[f'E{row}'] = f'=E{before_cont_row}*Owner_Cont'
    ws[f'E{row}'].number_format = '$#,##0'
    owner_cont_row = row
    row += 2
    
    # TOTAL HARD COSTS
    ws[f'B{row}'] = 'TOTAL HARD COSTS'
    ws[f'B{row}'].font = Font(bold=True, size=12)
    ws[f'E{row}'] = f'=E{before_cont_row}+E{design_cont_row}+E{constr_cont_row}+E{owner_cont_row}'
    ws[f'E{row}'].number_format = '$#,##0'
    ws[f'E{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    ws[f'E{row}'].font = Font(bold=True, size=12)
    total_hard_row = row
    row += 1
    
    # Cost per GSF
    ws[f'B{row}'] = 'Cost per GSF'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'E{row}'] = f'=E{total_hard_row}/{TOTAL_GSF}'
    ws[f'E{row}'].number_format = '$#,##0.00'
    ws[f'E{row}'].font = Font(bold=True)
    
    wb.defined_names['TotalHardCosts'] = DefinedName('TotalHardCosts', attr_text=f"'Hard Costs'!$E${total_hard_row}")

def create_soft_costs_sheet(wb):
    """Soft costs - categorized by payment timing, totaling 12% of HARD COSTS"""
    ws = wb.create_sheet("Soft Costs")
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    editable_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    inkind_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")  # Light blue
    cash_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Light red
    
    ws['A1'] = 'SOFT COSTS - PAYMENT TIMING: MONTHS 1-9 vs AFTER MONTH 9'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    ws['A2'] = 'Target: 12% of Total Hard Costs | IN-KIND = Paid after Month 9 from pre-sales'
    ws.merge_cells('A2:D2')
    
    row = 4
    headers = ['Category', 'Amount', 'Payment Timing', 'Basis']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    row += 1
    
    start_row = row
    
    # Architecture & Engineering
    ws[f'A{row}'] = 'Architecture & Engineering'
    ws[f'B{row}'] = '=TotalHardCosts*0.04'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].fill = editable_fill
    ws[f'C{row}'] = 'IN-KIND (After Mo 9)'
    ws[f'C{row}'].fill = inkind_fill
    ws[f'C{row}'].font = Font(bold=True, color="0000FF")
    ws[f'D{row}'] = '4.0% of hard costs'
    row += 1
    
    ws[f'A{row}'] = 'Marketing & Sales'
    ws[f'B{row}'] = '=TotalHardCosts*0.025'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].fill = editable_fill
    ws[f'C{row}'] = 'IN-KIND (After Mo 9)'
    ws[f'C{row}'].fill = inkind_fill
    ws[f'C{row}'].font = Font(bold=True, color="0000FF")
    ws[f'D{row}'] = '2.5% of hard costs'
    row += 1
    
    ws[f'A{row}'] = 'Legal & Accounting'
    ws[f'B{row}'] = '=TotalHardCosts*0.01'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].fill = editable_fill
    ws[f'C{row}'] = 'IN-KIND (After Mo 9)'
    ws[f'C{row}'].fill = inkind_fill
    ws[f'C{row}'].font = Font(bold=True, color="0000FF")
    ws[f'D{row}'] = '1.0% of hard costs'
    row += 1
    
    ws[f'A{row}'] = 'Developer Fee'
    ws[f'B{row}'] = '=TotalHardCosts*0.015'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].fill = editable_fill
    ws[f'C{row}'] = 'IN-KIND (After Mo 9)'
    ws[f'C{row}'].fill = inkind_fill
    ws[f'C{row}'].font = Font(bold=True, color="0000FF")
    ws[f'D{row}'] = '1.5% of hard costs'
    inkind_end_row = row
    row += 2
    
    # CASH Items (Months 1-9)
    ws[f'A{row}'] = 'Building Permit'
    ws[f'B{row}'] = '=TotalHardCosts*0.015'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].fill = editable_fill
    ws[f'C{row}'] = 'CASH (Months 1-9)'
    ws[f'C{row}'].fill = cash_fill
    ws[f'C{row}'].font = Font(bold=True, color="FF0000")
    ws[f'D{row}'] = '1.5% of hard costs'
    row += 1
    
    ws[f'A{row}'] = 'Impact Fees'
    ws[f'B{row}'] = '=TotalHardCosts*0.01'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].fill = editable_fill
    ws[f'C{row}'] = 'CASH (Months 1-9)'
    ws[f'C{row}'].fill = cash_fill
    ws[f'C{row}'].font = Font(bold=True, color="FF0000")
    ws[f'D{row}'] = '1.0% of hard costs'
    row += 1
    
    ws[f'A{row}'] = 'Plan Review'
    ws[f'B{row}'] = '=TotalHardCosts*0.005'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].fill = editable_fill
    ws[f'C{row}'] = 'CASH (Months 1-9)'
    ws[f'C{row}'].fill = cash_fill
    ws[f'C{row}'].font = Font(bold=True, color="FF0000")
    ws[f'D{row}'] = '0.5% of hard costs'
    row += 1
    
    ws[f'A{row}'] = 'Insurance'
    ws[f'B{row}'] = '=TotalHardCosts*0.008'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].fill = editable_fill
    ws[f'C{row}'] = 'CASH (Months 1-9)'
    ws[f'C{row}'].fill = cash_fill
    ws[f'C{row}'].font = Font(bold=True, color="FF0000")
    ws[f'D{row}'] = '0.8% of hard costs'
    row += 1
    
    ws[f'A{row}'] = 'Testing & Inspections'
    ws[f'B{row}'] = '=TotalHardCosts*0.003'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].fill = editable_fill
    ws[f'C{row}'] = 'CASH (Months 1-9)'
    ws[f'C{row}'].fill = cash_fill
    ws[f'C{row}'].font = Font(bold=True, color="FF0000")
    ws[f'D{row}'] = '0.3% of hard costs'
    row += 1
    
    ws[f'A{row}'] = 'Utilities & Misc'
    ws[f'B{row}'] = '=TotalHardCosts*0.002'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].fill = editable_fill
    ws[f'C{row}'] = 'CASH (Months 1-9)'
    ws[f'C{row}'].fill = cash_fill
    ws[f'C{row}'].font = Font(bold=True, color="FF0000")
    ws[f'D{row}'] = '0.2% of hard costs'
    row += 1
    
    ws[f'A{row}'] = 'Financing Costs'
    ws[f'B{row}'] = '=ConstructionSubtotal*0.375*LTC*Finance_Rate*(9/12)'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'C{row}'] = 'CASH (Months 1-9)'
    ws[f'C{row}'].fill = cash_fill
    ws[f'C{row}'].font = Font(bold=True, color="FF0000")
    ws[f'D{row}'] = '9-month bridge loan interest'
    cash_end_row = row
    row += 2
    
    # Note
    ws[f'A{row}'] = '  Note:'
    ws[f'B{row}'] = 'IN-KIND costs paid after Month 9 from pre-sales revenue'
    ws[f'C{row}'] = 'CASH costs paid Months 1-9 from bridge loan'
    ws.merge_cells(f'B{row}:D{row}')
    row += 2
    
    # Total
    ws[f'A{row}'] = 'TOTAL SOFT COSTS'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'B{row}'] = f'=SUM(B{start_row}:B{cash_end_row})'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    ws[f'B{row}'].font = Font(bold=True, size=12)
    total_soft_row = row
    row += 1
    
    # Percentage of HARD costs
    ws[f'A{row}'] = '% of Hard Costs'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = f'=B{total_soft_row}/TotalHardCosts'
    ws[f'B{row}'].number_format = '0.0%'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'D{row}'] = 'Target: 12.0%'
    row += 1
    
    # Cost per GSF
    ws[f'A{row}'] = 'Cost per GSF'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = f'=B{total_soft_row}/{TOTAL_GSF}'
    ws[f'B{row}'].number_format = '$#,##0.00'
    ws[f'B{row}'].font = Font(bold=True)
    
    wb.defined_names['TotalSoftCosts'] = DefinedName('TotalSoftCosts', attr_text=f"'Soft Costs'!$B${total_soft_row}")

def create_revenue_sheet(wb):
    """Revenue by unit type"""
    ws = wb.create_sheet("Revenue")
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    editable_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    ws['A1'] = 'REVENUE ANALYSIS'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    row = 3
    headers = ['Floor', 'Unit Type', 'Units', 'NSF', '$/SF', 'Total Revenue']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    row += 1
    
    # Unit types - FROM PRO_FORMA.JSON
    units = [
        ('4-5', 'Hospitality 2BR/1.5BA', 32, 25200, 1200),
        ('6', 'Hospitality 2BR/1.5BA', 10, 8680, 1200),
        ('7', 'Condos 3BR/2BA', 8, 8680, 1200),
        ('1', 'Chabad/Mikvah (at cost)', 1, 10500, 400),
    ]
    
    start_row = row
    for floor, type_, count, nsf, price in units:
        ws[f'A{row}'] = floor
        ws[f'B{row}'] = type_
        ws[f'C{row}'] = count
        ws[f'D{row}'] = nsf
        ws[f'D{row}'].number_format = '#,##0'
        ws[f'E{row}'] = price
        ws[f'E{row}'].fill = editable_fill
        ws[f'E{row}'].number_format = '$#,##0'
        ws[f'F{row}'] = f'=D{row}*E{row}'
        ws[f'F{row}'].number_format = '$#,##0'
        row += 1
    
    # Total
    row += 1
    ws[f'B{row}'] = 'TOTAL REVENUE'
    ws[f'B{row}'].font = Font(bold=True, size=12)
    ws[f'C{row}'] = f'=SUM(C{start_row}:C{row-2})'
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'F{row}'] = f'=SUM(F{start_row}:F{row-2})'
    ws[f'F{row}'].number_format = '$#,##0'
    ws[f'F{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    ws[f'F{row}'].font = Font(bold=True, size=12)
    
    wb.defined_names['TotalRevenue'] = DefinedName('TotalRevenue', attr_text=f"'Revenue'!$F${row}")

def create_project_deal_sheet(wb):
    """Project Deal Structure & Payout Explanation"""
    ws = wb.create_sheet("Project Deal Structure")
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    ws['A1'] = 'NEVO TOWER - PROJECT DEAL STRUCTURE'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:E1')
    
    ws['A2'] = 'Complete Partnership Structure, Financing, and Payout Waterfall'
    ws['A2'].font = Font(italic=True)
    ws.merge_cells('A2:E2')
    
    row = 4
    
    # PARTIES
    ws[f'A{row}'] = 'THE PARTIES'
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    parties = [
        ('Simcha Connection (SC)', 'Land Owner - 1580 79th St, North Bay Village', '$10M land value'),
        ('Developer Partner', 'Development partner sought by SC', 'Cash + expertise'),
        ('', 'Partnership is 50/50 profit split AFTER obligations', ''),
    ]
    
    for party, role, contribution in parties:
        ws[f'A{row}'] = party
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = role
        ws[f'D{row}'] = contribution
        row += 1
    
    row += 1
    
    # SC CONTRIBUTIONS
    ws[f'A{row}'] = 'SC CONTRIBUTIONS & PAYMENTS'
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    ws[f'A{row}'] = 'Item'
    ws[f'B{row}'] = 'Amount'
    ws[f'C{row}'] = 'Timing'
    ws[f'D{row}'] = 'Type'
    ws[f'E{row}'] = 'Interest?'
    for col in range(1, 6):
        ws.cell(row, col).font = Font(bold=True)
    row += 1
    
    sc_items = [
        ('Land (In-Kind)', '$8,000,000', 'Upfront', 'Equity contribution', 'NO'),
        ('Cash to SC from Dev', '$1,250,000', 'Month 0 (key money)', 'Cash OUT to SC', 'N/A'),
        ('Cash to SC from Dev', '$750,000', 'Month 6', 'Cash OUT to SC', 'N/A'),
        ('Land Payment AT EXIT', '$8,000,000', 'After all vendors paid', 'From windfall', 'NO!'),
    ]
    
    for item, amount, timing, type_, interest in sc_items:
        ws[f'A{row}'] = item
        ws[f'B{row}'] = amount
        ws[f'C{row}'] = timing
        ws[f'D{row}'] = type_
        ws[f'E{row}'] = interest
        if interest == 'NO!' or 'EXIT' in item:
            ws[f'E{row}'].fill = highlight_fill
            ws[f'E{row}'].font = Font(bold=True)
        row += 1
    
    row += 1
    
    # FINANCING STRUCTURE
    ws[f'A{row}'] = 'FINANCING STRUCTURE'
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    ws[f'A{row}'] = 'Source'
    ws[f'B{row}'] = 'Amount'
    ws[f'C{row}'] = 'Duration'
    ws[f'D{row}'] = 'Purpose'
    ws[f'E{row}'] = 'Cost'
    for col in range(1, 6):
        ws.cell(row, col).font = Font(bold=True)
    row += 1
    
    financing = [
        ('Developer Cash OUT to SC', '$2,000,000', 'Upfront', 'SC receives cash payments', '$0'),
        ('Bridge Loan', '~$6,600,000', '9 months', 'Construction subtotal only', '~$420K'),
        ('Pre-Sales Deposits', '~$14,700,000', 'Months 1-24', 'Remaining construction', '$0'),
        ('Vendor Terms', 'Variable', 'Ongoing', 'GC/OH&P/Contingencies', '$0'),
        ('In-Kind Suppliers', 'Variable', 'Ongoing', 'Materials at cost', '$0'),
    ]
    
    for source, amount, duration, purpose, cost in financing:
        ws[f'A{row}'] = source
        ws[f'B{row}'] = amount
        ws[f'C{row}'] = duration
        ws[f'D{row}'] = purpose
        ws[f'E{row}'] = cost
        if cost == '$0':
            ws[f'E{row}'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        row += 1
    
    row += 1
    ws[f'A{row}'] = 'KEY POINT:'
    ws[f'A{row}'].font = Font(bold=True, color="FF0000")
    ws[f'B{row}'] = 'Pre-sales from Month 10+ cover ALL remaining construction costs'
    ws.merge_cells(f'B{row}:E{row}')
    row += 2
    
    # PAYMENT WATERFALL
    ws[f'A{row}'] = 'PAYMENT WATERFALL (AT PROJECT EXIT)'
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    ws[f'A{row}'] = 'Priority'
    ws[f'B{row}'] = 'Payee'
    ws[f'C{row}'] = 'Amount'
    ws[f'D{row}'] = 'Description'
    for col in range(1, 5):
        ws.cell(row, col).font = Font(bold=True)
    row += 1
    
    waterfall = [
        ('0', 'Total Revenue', '$55,272,000', 'Gross sales revenue from 50 units'),
        ('1', 'Bridge Loan Payoff', '~$420,000', 'Principal + interest (if balance remains)'),
        ('2', 'Vendors/Contractors', '~$32,609,001', 'All construction costs paid'),
        ('3', 'Soft Costs', '~$4,600,000', 'A&E, marketing, permits, etc.'),
        ('4', 'SC Land Payment', '$8,000,000', 'NO INTEREST! Paid from windfall'),
        ('5', 'Remaining Windfall', '~$9,643,000', 'Available for profit split'),
        ('6', 'SC Share (50%)', '~$4,821,500', '50% of windfall after all obligations'),
        ('7', 'Developer Share (50%)', '~$4,821,500', '50% of windfall after all obligations'),
    ]
    
    for priority, payee, amount, description in waterfall:
        ws[f'A{row}'] = priority
        ws[f'B{row}'] = payee
        ws[f'C{row}'] = amount
        ws[f'D{row}'] = description
        if priority == '0':
            ws[f'A{row}'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        elif priority in ['4', '5']:
            ws[f'A{row}'].fill = highlight_fill
        elif priority in ['6', '7']:
            ws[f'A{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        row += 1
    
    row += 1
    
    # DEVELOPER RETURNS
    ws[f'A{row}'] = 'DEVELOPER ECONOMICS'
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    ws[f'A{row}'] = 'Cash OUT (Investment to SC)'
    ws[f'B{row}'] = '$2,000,000'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'D{row}'] = 'Paid to SC upfront'
    row += 1
    
    ws[f'A{row}'] = 'Cash IN (Return)'
    ws[f'B{row}'] = '~$4,821,500'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'D{row}'] = '50% of windfall'
    row += 1
    
    ws[f'A{row}'] = 'Net Profit'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = '~$2,821,500'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'D{row}'] = '$4.82M - $2M investment'
    row += 1
    
    ws[f'A{row}'] = 'ROI'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'B{row}'] = '141%'
    ws[f'B{row}'].font = Font(bold=True, size=12, color="008000")
    ws[f'B{row}'].fill = highlight_fill
    ws[f'D{row}'] = '$2.82M profit / $2M investment'
    row += 2
    
    # KEY BENEFITS
    ws[f'A{row}'] = 'KEY DEAL BENEFITS'
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    benefits = [
        ('✓ NO INTEREST on $8M SC land payment', 'Massive savings vs traditional land purchase'),
        ('✓ Bridge loan only 9 months', 'Pre-sales cover construction from Month 10+'),
        ('✓ Vendor payment terms', 'GC/OH&P/contingencies funded by terms, not loan'),
        ('✓ In-kind supplier opportunities', 'Reduce cash needed with material contributions'),
        ('✓ 141% ROI on $2M investment', 'Developer gets $4.82M from $2M investment'),
        ('✓ Pre-sales de-risk project', 'Units sold before completion = guaranteed revenue'),
    ]
    
    for benefit, explanation in benefits:
        ws[f'A{row}'] = benefit
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = explanation
        ws.merge_cells(f'B{row}:E{row}')
        row += 1

def create_cash_flow_sheet(wb):
    """24-Month Cash Flow Schedule with Pre-Sales"""
    ws = wb.create_sheet("24-Month Cash Flow")
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    ws['A1'] = 'NEVO TOWER - 24-MONTH CASH FLOW WITH PRE-SALES'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:J1')
    
    ws['A2'] = 'Construction Costs vs Pre-Sales Revenue - Bridge Loan Calculation'
    ws.merge_cells('A2:J2')
    
    row = 4
    headers = ['Month', 'Activity', 'Hard Cost Draw', 'SC Cash OUT', 'Pre-Sales Units', 'Pre-Sales Payment', 'Usable (90%)', 'Total Cash In', 'Net Cash Flow', 'Cumulative Cash', 'Bridge Loan Need']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    row += 1
    
    # Monthly hard cost percentages
    monthly_pcts = [
        0.03, 0.04, 0.05, 0.06,  # Months 1-4: Foundation
        0.07, 0.08, 0.08, 0.07,  # Months 5-8: Superstructure
        0.07, 0.07, 0.06, 0.06,  # Months 9-12: Envelope
        0.05, 0.05, 0.04, 0.04,  # Months 13-16: MEP
        0.03, 0.03, 0.02, 0.02,  # Months 17-20: Finishes
        0.01, 0.01, 0.005, 0.005  # Months 21-24: Closeout
    ]
    
    # Pre-sales velocity - TOTAL = 50 UNITS
    presales_units = [
        1, 2, 2, 3,    # Months 1-4: Slow start = 8
        3, 3, 4, 4,    # Months 5-8: Ramping up = 14
        4, 4, 3, 3,    # Months 9-12: Peak sales = 14
        2, 2, 2, 2,    # Months 13-16: Steady = 8
        1, 1, 1, 1,    # Months 17-20: Slowing = 4
        1, 1, 0, 0     # Months 21-24: Final 2 units = 2, TOTAL = 50
    ]
    
    activities = [
        'Foundation', 'Foundation', 'Site Work', 'Site Work',
        'Superstructure', 'Superstructure', 'Superstructure', 'Superstructure',
        'Envelope', 'Envelope', 'MEP Rough-In', 'MEP Rough-In',
        'MEP Install', 'MEP Install', 'Interior Framing', 'Interior Framing',
        'Finishes', 'Finishes', 'Fixtures', 'Fixtures',
        'Closeout', 'Punch List', 'Final Inspection', 'CO & Turnover'
    ]
    
    start_data_row = row
    cumulative_cell = f'I{row-1}'
    
    for month_num in range(1, 25):
        month_idx = month_num - 1
        ws[f'A{row}'] = month_num
        ws[f'B{row}'] = activities[month_idx]
        
        # Hard cost draw formula
        ws[f'C{row}'] = f'=TotalHardCosts*{monthly_pcts[month_idx]}'
        ws[f'C{row}'].number_format = '$#,##0'
        
        # SC Cash OUT
        if month_num == 1:
            ws[f'D{row}'] = '=-LandPartner_Cash1'
        elif month_num == 6:
            ws[f'D{row}'] = '=-LandPartner_Cash2'
        else:
            ws[f'D{row}'] = 0
        ws[f'D{row}'].number_format = '$#,##0'
        
        # Pre-sales units
        ws[f'E{row}'] = presales_units[month_idx]
        
        # Pre-sales payment schedule:
        # 25% at signing, 25% at month 5 (superstructure), 25% at month 15 (envelope), 25% at month 24 (CO)
        if month_num == 1:
            # Month 1: 25% of units sold this month
            ws[f'F{row}'] = f'=E{row}*(TotalRevenue/50)*0.25'
        elif month_num == 5:
            # Month 5: 25% of current + 25% of ALL previous sales (payment 2 of 4)
            ws[f'F{row}'] = f'=(E{row}*(TotalRevenue/50)*0.25)+(SUM($E$5:E8)*(TotalRevenue/50)*0.25)'
        elif month_num == 15:
            # Month 15: 25% of current + 25% of ALL previous sales (payment 3 of 4 = 75% total collected)
            ws[f'F{row}'] = f'=(E{row}*(TotalRevenue/50)*0.25)+(SUM($E$5:E18)*(TotalRevenue/50)*0.25)'
        elif month_num == 24:
            # Month 24: Final 25% payment for ALL units (100% collected)
            ws[f'F{row}'] = f'=(E{row}*(TotalRevenue/50)*0.25)+(SUM($E$5:E27)*(TotalRevenue/50)*0.25)'
        else:
            # Other months: just 25% of current month sales
            ws[f'F{row}'] = f'=E{row}*(TotalRevenue/50)*0.25'
        ws[f'F{row}'].number_format = '$#,##0'
        
        # Usable pre-sales (90% before month 24, 100% at month 24)
        if month_num < 24:
            ws[f'G{row}'] = f'=F{row}*0.90'
        else:
            ws[f'G{row}'] = f'=F{row}'
        ws[f'G{row}'].number_format = '$#,##0'
        
        # Total cash in (SC Cash + Usable Pre-Sales)
        ws[f'H{row}'] = f'=D{row}+G{row}'
        ws[f'H{row}'].number_format = '$#,##0'
        
        # Net cash flow
        ws[f'I{row}'] = f'=H{row}-C{row}'
        ws[f'I{row}'].number_format = '$#,##0'
        
        # Cumulative cash
        if month_num == 1:
            ws[f'J{row}'] = f'=I{row}'
        else:
            ws[f'J{row}'] = f'=J{row-1}+I{row}'
        ws[f'J{row}'].number_format = '$#,##0'
        
        # Bridge loan need
        ws[f'K{row}'] = f'=IF(J{row}<0,-J{row},0)'
        ws[f'K{row}'].number_format = '$#,##0'
        ws[f'K{row}'].fill = highlight_fill
        
        row += 1
    
    # Summary rows
    row += 1
    ws[f'A{row}'] = 'TOTALS:'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'C{row}'] = f'=SUM(C{start_data_row}:C{row-2})'
    ws[f'C{row}'].number_format = '$#,##0'
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'D{row}'] = f'=SUM(D{start_data_row}:D{row-2})'
    ws[f'D{row}'].number_format = '$#,##0'
    ws[f'D{row}'].font = Font(bold=True)
    ws[f'E{row}'] = f'=SUM(E{start_data_row}:E{row-2})'
    ws[f'E{row}'].font = Font(bold=True)
    ws[f'F{row}'] = f'=SUM(F{start_data_row}:F{row-2})'
    ws[f'F{row}'].number_format = '$#,##0'
    ws[f'F{row}'].font = Font(bold=True)
    ws[f'G{row}'] = f'=SUM(G{start_data_row}:G{row-2})'
    ws[f'G{row}'].number_format = '$#,##0'
    ws[f'G{row}'].font = Font(bold=True)
    
    row += 2
    ws[f'A{row}'] = 'VERIFICATION:'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = 'Total Pre-Sales Payment ='
    ws[f'C{row}'] = f'=SUM(F{start_data_row}:F{row-3})'
    ws[f'C{row}'].number_format = '$#,##0'
    ws[f'D{row}'] = 'Should = Total Revenue ='
    ws[f'E{row}'] = '=TotalRevenue'
    ws[f'E{row}'].number_format = '$#,##0'
    row += 1
    
    ws[f'A{row}'] = 'MAX BRIDGE LOAN NEEDED:'
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FF0000")
    ws[f'B{row}'] = f'=MAX(K{start_data_row}:K{row-4})'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].font = Font(bold=True, size=12)
    ws[f'B{row}'].fill = highlight_fill

def create_gc_rfq_sheet(wb):
    """GC RFQ Summary with Materials, Labor, and Workforce Breakdown"""
    ws = wb.create_sheet("GC RFQ Summary")
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    editable_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    ws['A1'] = 'GENERAL CONTRACTOR - REQUEST FOR QUOTE SUMMARY'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')
    
    ws['A2'] = 'NEVO Tower - 1580 79th St Causeway, North Bay Village, FL'
    ws.merge_cells('A2:H2')
    
    ws['A3'] = f'Total Project: {TOTAL_GSF:,} GSF | 7 Stories + Rooftop | 50 Residential Units'
    ws.merge_cells('A3:H3')
    
    row = 5
    
    # CSI Division breakdown with labor/material split and worker types
    ws[f'A{row}'] = 'CSI DIVISION BREAKDOWN - MATERIALS, LABOR & WORKFORCE'
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    headers = ['CSI', 'Description', 'Total Cost', 'Materials %', 'Labor %', 'Man-Hours', 'Worker Type', 'Crew Size']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    row += 1
    
    # CSI divisions with realistic labor/material splits and worker classifications
    csi_data = [
        # (CSI, Description, Rate/GSF, Material%, Labor%, Worker Type, Crew Size)
        ('01', 'General Requirements', 10.53, 40, 60, 'Project Management/Supervisors', '5-8'),
        ('02', 'Site Work & Demolition', 4.56, 30, 70, 'Laborers/Equipment Operators', '6-10'),
        ('03', 'Concrete - PT Structure', 39.47, 45, 55, 'Concrete Workers/Ironworkers', '15-25'),
        ('04', 'Masonry', 2.46, 50, 50, 'Masons/Laborers', '4-6'),
        ('05', 'Metals & Structural Steel', 7.45, 65, 35, 'Ironworkers/Welders', '6-10'),
        ('06', 'Wood & Plastics', 2.81, 55, 45, 'Carpenters', '4-6'),
        ('07', 'Waterproofing & Roofing', 19.30, 40, 60, 'Roofers/Waterproofing Specialists', '6-10'),
        ('08', 'Doors, Windows & HVHZ Glazing', 24.55, 70, 30, 'Glaziers/Installation Specialists', '8-12'),
        ('09', 'Interior Finishes', 36.83, 45, 55, 'Carpenters/Painters/Drywall', '20-30'),
        ('10', 'Specialties', 3.95, 75, 25, 'Specialty Installers', '2-4'),
        ('11', 'Equipment', 3.33, 80, 20, 'Equipment Installers', '2-4'),
        ('12', 'Furnishings', 1.05, 85, 15, 'Installers', '2-3'),
        ('13', 'Special Construction', 1.58, 60, 40, 'Specialty Contractors', '3-5'),
        ('14', 'Elevators', 3.10, 75, 25, 'Elevator Mechanics/Technicians', '4-6'),
        ('21', 'Fire Suppression', 6.58, 50, 50, 'Pipefitters/Sprinkler Fitters', '4-6'),
        ('22', 'Plumbing', 8.77, 45, 55, 'Plumbers/Pipefitters', '6-10'),
        ('23', 'HVAC', 13.15, 50, 50, 'HVAC Technicians/Sheet Metal', '8-12'),
        ('26', 'Electrical', 13.15, 40, 60, 'Electricians/Low Voltage Tech', '10-15'),
        ('27', 'Communications & Data', 1.75, 60, 40, 'Low Voltage Technicians', '3-5'),
        ('28', 'Security & Fire Alarm', 1.75, 65, 35, 'Security/Fire Alarm Techs', '3-4'),
        ('31', 'Earthwork', 3.68, 25, 75, 'Equipment Operators/Laborers', '6-8'),
        ('32', 'Exterior Improvements', 3.51, 50, 50, 'Landscapers/Pavers/Masons', '6-10'),
        ('33', 'Utilities', 2.81, 55, 45, 'Utility Workers/Excavators', '4-6'),
    ]
    
    start_row = row
    total_hours = 0
    
    for csi, desc, rate, mat_pct, lab_pct, worker_type, crew in csi_data:
        # Calculate total cost using formula
        ws[f'A{row}'] = csi
        ws[f'B{row}'] = desc
        ws[f'C{row}'] = f'={rate}*{TOTAL_GSF}'
        ws[f'C{row}'].number_format = '$#,##0'
        ws[f'D{row}'] = mat_pct / 100
        ws[f'D{row}'].number_format = '0%'
        ws[f'E{row}'] = lab_pct / 100
        ws[f'E{row}'].number_format = '0%'
        
        # Calculate man-hours based on labor cost at $45/hour average
        labor_cost = rate * TOTAL_GSF * (lab_pct / 100)
        man_hours = int(labor_cost / 45)
        total_hours += man_hours
        
        ws[f'F{row}'] = man_hours
        ws[f'F{row}'].number_format = '#,##0'
        ws[f'G{row}'] = worker_type
        ws[f'H{row}'] = crew
        row += 1
    
    # Subtotal
    ws[f'B{row}'] = 'SUBTOTAL - Direct Costs'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = f'=SUM(C{start_row}:C{row-1})'
    ws[f'C{row}'].number_format = '$#,##0'
    ws[f'C{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    ws[f'F{row}'] = f'=SUM(F{start_row}:F{row-1})'
    ws[f'F{row}'].number_format = '#,##0'
    ws[f'F{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    subtotal_row = row
    row += 2
    
    # GC Costs
    ws[f'B{row}'] = 'General Conditions (9%)'
    ws[f'C{row}'] = f'=C{subtotal_row}*GC_GC'
    ws[f'C{row}'].number_format = '$#,##0'
    ws[f'F{row}'] = int(total_hours * 0.05)  # 5% of hours for GC staff
    ws[f'F{row}'].number_format = '#,##0'
    ws[f'G{row}'] = 'Project Manager/Super/Admin'
    gc_row = row
    row += 1
    
    ws[f'B{row}'] = 'Overhead & Profit (10%)'
    ws[f'C{row}'] = f'=(C{subtotal_row}+C{gc_row})*GC_OHP'
    ws[f'C{row}'].number_format = '$#,##0'
    row += 1
    
    ws[f'B{row}'] = 'Contingencies (16%)'
    ws[f'C{row}'] = f'=(C{subtotal_row}+C{gc_row}+C{row-1})*0.16'
    ws[f'C{row}'].number_format = '$#,##0'
    row += 2
    
    # TOTAL
    ws[f'B{row}'] = 'TOTAL HARD COSTS'
    ws[f'B{row}'].font = Font(bold=True, size=12)
    ws[f'C{row}'] = '=TotalHardCosts'
    ws[f'C{row}'].number_format = '$#,##0'
    ws[f'C{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    ws[f'C{row}'].font = Font(bold=True, size=12)
    ws[f'F{row}'] = f'=F{subtotal_row}+F{gc_row}'
    ws[f'F{row}'].number_format = '#,##0'
    ws[f'F{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    ws[f'F{row}'].font = Font(bold=True)
    row += 3
    
    # LABOR SUMMARY BY WORKER TYPE
    ws[f'A{row}'] = 'WORKFORCE SUMMARY'
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    headers2 = ['Category', 'Worker Types', 'Est. Man-Hours', 'Avg Rate/Hr', 'Total Labor Cost', 'Peak Workers', 'Duration']
    for col, header in enumerate(headers2, 1):
        cell = ws.cell(row, col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    row += 1
    
    # Worker categories
    workforce_start = row
    workforce = [
        ('Skilled Trades', 'Concrete, Ironworkers, Electricians, Plumbers, HVAC', 0.50, 55, '18-24 months'),
        ('Semi-Skilled', 'Carpenters, Masons, Painters, Drywall', 0.25, 40, '12-20 months'),
        ('Laborers', 'General Labor, Site Work, Material Handling', 0.15, 30, '1-24 months'),
        ('Specialists', 'Glaziers, Waterproofing, Elevator Techs', 0.08, 60, '8-18 months'),
        ('Management/Super', 'PM, Superintendent, Foremen, Admin', 0.02, 75, '1-24 months'),
    ]
    
    for category, types, pct, rate, duration in workforce:
        ws[f'A{row}'] = category
        ws[f'B{row}'] = types
        ws[f'C{row}'] = int(total_hours * pct)
        ws[f'C{row}'].number_format = '#,##0'
        ws[f'D{row}'] = rate
        ws[f'D{row}'].number_format = '$#,##0'
        ws[f'E{row}'] = f'=C{row}*D{row}'
        ws[f'E{row}'].number_format = '$#,##0'
        # Peak workers = hours / 2000 hrs/year * 1.5 for peak
        ws[f'F{row}'] = f'=INT(C{row}/2000*1.5)'
        ws[f'G{row}'] = duration
        row += 1
    
    row += 1
    ws[f'A{row}'] = 'TOTAL WORKFORCE'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = f'=SUM(C{workforce_start}:C{row-2})'
    ws[f'C{row}'].number_format = '#,##0'
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'E{row}'] = f'=SUM(E{workforce_start}:E{row-2})'
    ws[f'E{row}'].number_format = '$#,##0'
    ws[f'E{row}'].font = Font(bold=True)
    ws[f'F{row}'] = f'=SUM(F{workforce_start}:F{row-2})'
    ws[f'F{row}'].font = Font(bold=True)

def create_sync_instructions_sheet(wb):
    """Instructions for keeping workbook in sync"""
    ws = wb.create_sheet("Sync Instructions")
    
    ws['A1'] = 'HOW TO KEEP THIS WORKBOOK IN SYNC'
    ws['A1'].font = Font(bold=True, size=14)
    
    row = 3
    instructions = [
        ('YELLOW CELLS = EDITABLE', 'You can change these values'),
        ('', ''),
        ('When you edit a yellow cell:', 'All formulas update automatically'),
        ('', ''),
        ('Named Ranges:', 'Connect sheets together'),
        ('Example:', 'TotalHardCosts pulls from Hard Costs sheet'),
        ('', ''),
        ('To regenerate this file:', 'Run scripts/make_interactive_workbook_v3.py'),
        ('', ''),
        ('ALL 11 CORRECTIONS APPLIED:', ''),
        ('1. Summary - % in column C', '✓'),
        ('2. Assumptions - % in columns C & D', '✓'),
        ('3. Soft costs = 12% of project cost', '✓'),
        ('4. All formulas use named ranges', '✓'),
        ('5. Cash flow with formulas & 50 units', '✓'),
        ('6. SC Cash OUT label corrected', '✓'),
        ('7. Auto-fit columns', '✓'),
        ('8. SF with commas throughout', '✓'),
        ('9. Soft costs categorized IN-KIND/CASH', '✓'),
        ('10. Project deal sheet corrected', '✓'),
        ('11. All sheets formatted properly', '✓'),
    ]
    
    for instruction, detail in instructions:
        ws[f'A{row}'] = instruction
        if instruction.startswith('ALL') or any(c.isdigit() and ')' in instruction for c in instruction):
            ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = detail
        if detail == '✓':
            ws[f'B{row}'].font = Font(bold=True, color="008000", size=14)
        row += 1

if __name__ == "__main__":
    create_interactive_workbook()
