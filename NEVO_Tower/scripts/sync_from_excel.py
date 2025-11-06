"""
Sync from Excel: Read edited workbook and regenerate all Python models
Two-way sync: Excel → Python JSON files → Reports
"""

import openpyxl
import json
from datetime import datetime
from pathlib import Path

EXCEL_FILE = 'outputs/NEVO_Interactive_Budget.xlsx'

def sync_from_excel():
    """Read Excel workbook and update all Python data files"""
    
    print("=" * 70)
    print("NEVO TOWER - SYNCING FROM EXCEL")
    print("=" * 70)
    print(f"Reading: {EXCEL_FILE}\n")
    
    # Try to use Excel COM to calculate formulas first
    try:
        import win32com.client
        import os
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        abs_path = os.path.abspath(EXCEL_FILE)
        workbook = excel.Workbooks.Open(abs_path)
        workbook.Save()
        workbook.Close()
        excel.Quit()
        print("✓ Formulas calculated via Excel COM")
    except Exception as e:
        print(f"⚠ Could not use Excel COM: {e}")
        print("  Please save the Excel file in Excel first to calculate formulas.")
    
    # Load workbook with calculated values
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    
    # Extract all data
    assumptions = extract_assumptions(wb)
    hard_costs = extract_hard_costs(wb)
    soft_costs = extract_soft_costs(wb)
    revenue = extract_revenue(wb)
    summary = extract_summary(wb)
    
    # Save to JSON files
    save_data(assumptions, hard_costs, soft_costs, revenue, summary)
    
    # Regenerate reports
    regenerate_reports(assumptions, hard_costs, soft_costs, revenue, summary)
    
    print("\n" + "=" * 70)
    print("✅ SYNC COMPLETE")
    print("=" * 70)
    print("\nUpdated files:")
    print("  - data/hard_costs_detail.json")
    print("  - data/soft_costs_detail.json")
    print("  - data/pro_forma.json")
    print("  - outputs/UPDATED_SUMMARY.md")
    print("\nYou can now:")
    print("  1. Make more edits in Excel")
    print("  2. Re-run this script anytime")
    print("  3. Or regenerate the workbook: python scripts\\make_interactive_workbook.py")

def extract_assumptions(wb):
    """Extract editable assumptions"""
    ws = wb['Assumptions']
    
    assumptions = {
        'timestamp': datetime.now().isoformat(),
        'project': {},
        'costs': {}
    }
    
    # Project info (rows 5-10)
    assumptions['project']['name'] = ws['B5'].value
    assumptions['project']['address'] = ws['B6'].value
    assumptions['project']['total_gsf'] = ws['B7'].value
    assumptions['project']['total_nsf'] = ws['B8'].value
    assumptions['project']['residential_units'] = ws['B9'].value
    assumptions['project']['parking_spaces'] = ws['B10'].value
    
    # Cost assumptions (rows 15-25)
    assumptions['costs']['land_cost'] = ws['B15'].value
    assumptions['costs']['gc_ohp'] = ws['B16'].value
    assumptions['costs']['gc_general_conditions'] = ws['B17'].value
    assumptions['costs']['design_contingency'] = ws['B18'].value
    assumptions['costs']['construction_contingency'] = ws['B19'].value
    assumptions['costs']['owner_contingency'] = ws['B20'].value
    assumptions['costs']['ae_fee_rate'] = ws['B21'].value
    assumptions['costs']['marketing_rate'] = ws['B22'].value
    assumptions['costs']['financing_rate'] = ws['B23'].value
    assumptions['costs']['loan_to_cost'] = ws['B24'].value
    assumptions['costs']['construction_duration'] = ws['B25'].value
    
    print("✓ Assumptions extracted")
    return assumptions

def extract_hard_costs(wb):
    """Extract hard costs from CSI divisions"""
    ws = wb['Hard Costs']
    
    hard_costs = {
        'timestamp': datetime.now().isoformat(),
        'divisions': [],
        'subtotals': {},
        'total': 0
    }
    
    # CSI Divisions (rows 4-26)
    for row in range(4, 27):
        div = {
            'division': ws[f'A{row}'].value,
            'description': ws[f'B{row}'].value,
            'rate_per_gsf': ws[f'C{row}'].value,
            'gsf': ws[f'D{row}'].value,
            'total': ws[f'E{row}'].value
        }
        hard_costs['divisions'].append(div)
    
    # Extract calculated totals
    hard_costs['subtotals']['construction'] = ws['E27'].value
    hard_costs['subtotals']['general_conditions'] = ws['E29'].value
    hard_costs['subtotals']['ohp'] = ws['E30'].value
    hard_costs['subtotals']['before_contingency'] = ws['E32'].value
    hard_costs['subtotals']['design_contingency'] = ws['E34'].value
    hard_costs['subtotals']['construction_contingency'] = ws['E35'].value
    hard_costs['subtotals']['owner_contingency'] = ws['E36'].value
    hard_costs['total'] = ws['E38'].value
    
    # Convert to float if string
    if isinstance(hard_costs['total'], str):
        hard_costs['total'] = float(hard_costs['total'].replace('$', '').replace(',', ''))
    
    print(f"✓ Hard costs extracted: ${hard_costs['total']:,.0f}")
    return hard_costs

def extract_soft_costs(wb):
    """Extract soft costs"""
    ws = wb['Soft Costs']
    
    soft_costs = {
        'timestamp': datetime.now().isoformat(),
        'items': [],
        'total': 0
    }
    
    # Soft cost items (rows 4-14)
    items_map = [
        ('Architecture & Engineering', 4),
        ('Building Permit', 5),
        ('Impact Fees', 6),
        ('Plan Review', 7),
        ('Legal & Accounting', 8),
        ('Marketing & Sales', 9),
        ('Insurance', 10),
        ('Financing Costs', 11),
        ('Testing & Inspections', 12),
        ('Utilities & Misc', 13),
        ('Developer Fee', 14),
    ]
    
    for name, row in items_map:
        item = {
            'category': name,
            'amount': ws[f'B{row}'].value,
            'basis': ws[f'C{row}'].value
        }
        soft_costs['items'].append(item)
    
    soft_costs['total'] = ws['B16'].value
    
    # Convert to float if string
    if isinstance(soft_costs['total'], str):
        soft_costs['total'] = float(soft_costs['total'].replace('$', '').replace(',', ''))
    
    print(f"✓ Soft costs extracted: ${soft_costs['total']:,.0f}")
    return soft_costs

def extract_revenue(wb):
    """Extract revenue data"""
    ws = wb['Revenue']
    
    revenue = {
        'timestamp': datetime.now().isoformat(),
        'unit_types': [],
        'total': 0
    }
    
    # Unit types (rows 4-7)
    for row in range(4, 8):
        unit = {
            'floor': ws[f'A{row}'].value,
            'type': ws[f'B{row}'].value,
            'units': ws[f'C{row}'].value,
            'nsf': ws[f'D{row}'].value,
            'price_per_sf': ws[f'E{row}'].value,
            'total_revenue': ws[f'F{row}'].value
        }
        revenue['unit_types'].append(unit)
    
    revenue['total'] = ws['F9'].value
    
    # Convert to float if string
    if isinstance(revenue['total'], str):
        revenue['total'] = float(revenue['total'].replace('$', '').replace(',', ''))
    
    print(f"✓ Revenue extracted: ${revenue['total']:,.0f}")
    return revenue

def extract_summary(wb):
    """Extract summary metrics"""
    ws = wb['Summary']
    
    summary = {
        'timestamp': datetime.now().isoformat(),
        'costs': {
            'land': ws['B5'].value,
            'hard_costs': ws['B6'].value,
            'soft_costs': ws['B7'].value,
            'total': ws['B8'].value
        },
        'revenue': {
            'total': ws['B11'].value
        },
        'profitability': {
            'gross_profit': ws['B14'].value,
            'profit_margin': ws['B15'].value,
            'roi': ws['B16'].value
        },
        'metrics': {
            'cost_per_unit': ws['B18'].value,
            'cost_per_nsf': ws['B19'].value,
            'revenue_per_nsf': ws['B20'].value
        }
    }
    
    print(f"✓ Summary extracted")
    print(f"  Total Cost:   ${summary['costs']['total']:,.0f}")
    print(f"  Total Revenue: ${summary['revenue']['total']:,.0f}")
    print(f"  Profit:       ${summary['profitability']['gross_profit']:,.0f}")
    print(f"  Margin:       {summary['profitability']['profit_margin']:.1%}")
    
    return summary

def save_data(assumptions, hard_costs, soft_costs, revenue, summary):
    """Save all data to JSON files"""
    
    # Update hard_costs_detail.json
    with open('data/hard_costs_detail.json', 'w') as f:
        json.dump(hard_costs, f, indent=2)
    
    # Update soft_costs_detail.json
    with open('data/soft_costs_detail.json', 'w') as f:
        json.dump(soft_costs, f, indent=2)
    
    # Update pro_forma.json (combined financial model)
    pro_forma = {
        'timestamp': datetime.now().isoformat(),
        'assumptions': assumptions,
        'revenue': revenue,
        'costs': {
            'land': assumptions['costs']['land_cost'],
            'hard_costs': hard_costs,
            'soft_costs': soft_costs,
            'total': summary['costs']['total']
        },
        'profitability': summary['profitability'],
        'metrics': summary['metrics']
    }
    
    with open('data/pro_forma.json', 'w') as f:
        json.dump(pro_forma, f, indent=2)
    
    print("\n✓ JSON files updated")

def regenerate_reports(assumptions, hard_costs, soft_costs, revenue, summary):
    """Generate updated markdown reports"""
    
    report = f"""# NEVO TOWER - UPDATED FINANCIAL SUMMARY
*Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*
*Source: Excel workbook edits*

---

## PROJECT OVERVIEW

**{assumptions['project']['name']}**
{assumptions['project']['address']}

- Total GSF: {assumptions['project']['total_gsf']:,}
- Total NSF: {assumptions['project']['total_nsf']:,}
- Residential Units: {assumptions['project']['residential_units']}
- Parking Spaces: {assumptions['project']['parking_spaces']}

---

## DEVELOPMENT COSTS

### Land Acquisition
**${assumptions['costs']['land_cost']:,.0f}**

### Hard Costs (CSI MasterFormat)
"""

    # Add CSI divisions
    for div in hard_costs['divisions']:
        report += f"\n- **{div['division']}** {div['description']}: ${div['rate_per_gsf']:.2f}/GSF × {div['gsf']:,} = **${div['total']:,.0f}**"
    
    report += f"""

#### Hard Cost Subtotals
- Construction Subtotal: ${hard_costs['subtotals']['construction']:,.0f}
- General Conditions ({assumptions['costs']['gc_general_conditions']:.1%}): ${hard_costs['subtotals']['general_conditions']:,.0f}
- Overhead & Profit ({assumptions['costs']['gc_ohp']:.1%}): ${hard_costs['subtotals']['ohp']:,.0f}
- **Before Contingency: ${hard_costs['subtotals']['before_contingency']:,.0f}**

#### Contingencies
- Design ({assumptions['costs']['design_contingency']:.1%}): ${hard_costs['subtotals']['design_contingency']:,.0f}
- Construction ({assumptions['costs']['construction_contingency']:.1%}): ${hard_costs['subtotals']['construction_contingency']:,.0f}
- Owner ({assumptions['costs']['owner_contingency']:.1%}): ${hard_costs['subtotals']['owner_contingency']:,.0f}

### **TOTAL HARD COSTS: ${hard_costs['total']:,.0f}**

---

### Soft Costs
"""

    for item in soft_costs['items']:
        report += f"\n- {item['category']}: **${item['amount']:,.0f}**"
        if item['basis']:
            report += f" ({item['basis']})"
    
    report += f"""

### **TOTAL SOFT COSTS: ${soft_costs['total']:,.0f}**

---

## TOTAL PROJECT COST
# **${summary['costs']['total']:,.0f}**

---

## REVENUE

"""

    for unit in revenue['unit_types']:
        report += f"\n- **{unit['type']}** (Floor {unit['floor']}): {unit['units']} units × {unit['nsf']:,} NSF @ ${unit['price_per_sf']:,.0f}/SF = **${unit['total_revenue']:,.0f}**"
    
    report += f"""

## TOTAL REVENUE
# **${revenue['total']:,.0f}**

---

## PROFITABILITY ANALYSIS

| Metric | Value |
|--------|-------|
| **Gross Profit** | **${summary['profitability']['gross_profit']:,.0f}** |
| **Profit Margin** | **{summary['profitability']['profit_margin']:.1%}** |
| **Return on Cost** | **{summary['profitability']['roi']:.1%}** |

### Key Metrics

- **Cost per Unit**: ${summary['metrics']['cost_per_unit']:,.0f}
- **Cost per NSF**: ${summary['metrics']['cost_per_nsf']:,.2f}
- **Revenue per NSF**: ${summary['metrics']['revenue_per_nsf']:,.2f}

---

## FEASIBILITY ASSESSMENT

"""

    profit = summary['profitability']['gross_profit']
    margin = summary['profitability']['profit_margin']
    
    if profit > 0 and margin > 0.15:
        status = "✅ **FEASIBLE** - Strong profitability"
    elif profit > 0 and margin > 0.08:
        status = "⚠️ **MARGINAL** - Acceptable but tight margins"
    elif profit > 0:
        status = "⚠️ **MARGINAL** - Low profit margins, high risk"
    else:
        status = "❌ **NOT FEASIBLE** - Project underwater"
    
    report += f"{status}\n\n"
    
    if profit <= 0:
        breakeven_revenue = summary['costs']['total']
        current_nsf = assumptions['project']['total_nsf']
        breakeven_price = breakeven_revenue / current_nsf
        report += f"""### Required Changes for Breakeven
- **Increase revenue to**: ${breakeven_revenue:,.0f} (+{((breakeven_revenue/revenue['total'])-1)*100:.1f}%)
- **OR increase pricing to**: ${breakeven_price:,.0f}/SF (+{((breakeven_price/(revenue['total']/current_nsf))-1)*100:.1f}%)
- **OR reduce costs by**: ${-profit:,.0f} ({-profit/summary['costs']['total']*100:.1f}%)

"""

    report += f"""---

## EDIT WORKFLOW

1. **Open Excel**: `{EXCEL_FILE}`
2. **Edit yellow cells**: All rates, quantities, and assumptions
3. **Save the file**
4. **Run sync**: `python scripts\\sync_from_excel.py`
5. **Review this report**: Updates automatically

---

*This report was auto-generated from your Excel edits.*
"""

    # Save report
    output_file = 'outputs/UPDATED_SUMMARY.md'
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(report)
    
    print(f"✓ Report generated: {output_file}")

if __name__ == "__main__":
    sync_from_excel()
