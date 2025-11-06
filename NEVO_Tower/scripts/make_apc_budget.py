"""
Create Alternative NEVO Tower Budget using APC Filigree Precast System
Reduces concrete labor, formwork, and construction time
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.workbook.defined_name import DefinedName
from datetime import datetime
import json

# Load base data
with open('data/nevo_parsed.json', 'r') as f:
    project_data = json.load(f)

TOTAL_GSF = project_data['building_totals']['total_gsf']
TOTAL_NSF = project_data['building_totals']['total_nsf']
RESIDENTIAL_UNITS = project_data['building_totals']['residential_units']

def create_apc_workbook():
    """Create workbook with APC Filigree system pricing"""
    
    # Load existing workbook
    wb = openpyxl.load_workbook('outputs/NEVO_Interactive_Budget_V3.xlsx')
    
    print("Creating APC Filigree Alternative Budget...")
    
    # Create new Hard Costs sheet
    if 'Hard Costs - APC Filigree' in wb.sheetnames:
        del wb['Hard Costs - APC Filigree']
    
    create_apc_hard_costs_sheet(wb)
    
    # Create detailed RFQ
    if 'RFQ - APC Filigree Detail' in wb.sheetnames:
        del wb['RFQ - APC Filigree Detail']
    
    create_apc_rfq_detail(wb)
    
    # Save
    filename = 'outputs/NEVO_APC_Budget.xlsx'
    wb.save(filename)
    
    print(f"\n✅ APC Filigree Budget created: {filename}")
    return filename

def create_apc_hard_costs_sheet(wb):
    """Hard costs using APC Filigree Precast System"""
    ws = wb.create_sheet("Hard Costs - APC Filigree", 2)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    editable_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    savings_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    
    ws['A1'] = 'HARD COSTS - APC FILIGREE PRECAST SYSTEM'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    ws['A2'] = 'Aerial Precast Concrete Composite Deck System - Reduced Labor & Time'
    ws.merge_cells('A2:F2')
    
    row = 4
    headers = ['CSI', 'Description', 'Rate ($/GSF)', 'GSF', 'Total Cost', 'vs Traditional']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    row += 1
    
    # CSI Divisions with APC Filigree pricing
    # CSI 03 significantly reduced, formwork eliminated
    csi_divisions = [
        ('01', 'General Requirements', 10.53, 0, 'Same'),
        ('02', 'Existing Conditions & Site Work', 4.56, 0, 'Same'),
        ('03', 'APC Filigree Precast System', 31.58, -7.89, '20% savings'),  # Was 39.47
        ('04', 'Masonry', 2.46, 0, 'Same'),
        ('05', 'Structural Steel (integrated)', 5.96, -1.49, '20% savings'),  # Was 7.45
        ('06', 'Wood & Plastics', 2.81, 0, 'Same'),
        ('07', 'Thermal & Moisture Protection', 19.30, 0, 'Same'),
        ('08', 'Openings (Doors, Windows, HVHZ)', 24.55, 0, 'Same'),
        ('09', 'Finishes', 36.83, 0, 'Same'),
        ('10', 'Specialties', 3.95, 0, 'Same'),
        ('11', 'Equipment', 3.33, 0, 'Same'),
        ('12', 'Furnishings', 1.05, 0, 'Same'),
        ('13', 'Special Construction', 1.58, 0, 'Same'),
        ('14', 'Conveying Equipment', 2.72, 0, 'Same'),
        ('21', 'Fire Suppression', 6.58, 0, 'Same'),
        ('22', 'Plumbing', 8.77, 0, 'Same'),
        ('23', 'HVAC', 13.15, 0, 'Same'),
        ('26', 'Electrical', 13.15, 0, 'Same'),
        ('27', 'Communications', 1.75, 0, 'Same'),
        ('28', 'Electronic Safety & Security', 1.75, 0, 'Same'),
        ('31', 'Earthwork', 3.68, 0, 'Same'),
        ('32', 'Exterior Improvements', 3.51, 0, 'Same'),
        ('33', 'Utilities', 2.81, 0, 'Same'),
    ]
    
    start_row = row
    traditional_total = 0
    apc_total = 0
    
    for div, desc, rate, savings, note in csi_divisions:
        ws[f'A{row}'] = div
        ws[f'B{row}'] = desc
        ws[f'C{row}'] = rate
        ws[f'C{row}'].fill = editable_fill
        ws[f'C{row}'].number_format = '$#,##0.00'
        ws[f'D{row}'] = TOTAL_GSF
        ws[f'D{row}'].number_format = '#,##0'
        ws[f'E{row}'] = f'=C{row}*D{row}'
        ws[f'E{row}'].number_format = '$#,##0'
        ws[f'F{row}'] = note
        if savings < 0:
            ws[f'F{row}'].fill = savings_fill
            ws[f'F{row}'].font = Font(bold=True, color="006100")
        
        traditional_total += (rate - savings) * TOTAL_GSF
        apc_total += rate * TOTAL_GSF
        row += 1
    
    # Subtotal
    ws[f'B{row}'] = 'SUBTOTAL - Construction'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'E{row}'] = f'=SUM(E{start_row}:E{row-1})'
    ws[f'E{row}'].number_format = '$#,##0'
    ws[f'E{row}'].fill = total_fill
    subtotal_row = row
    row += 2
    
    # Comparison to traditional
    ws[f'B{row}'] = 'Traditional PT System Cost'
    ws[f'E{row}'] = traditional_total
    ws[f'E{row}'].number_format = '$#,##0'
    ws[f'F{row}'] = 'For reference'
    trad_row = row
    row += 1
    
    ws[f'B{row}'] = 'APC Filigree System Cost'
    ws[f'E{row}'] = f'=E{subtotal_row}'
    ws[f'E{row}'].number_format = '$#,##0'
    ws[f'E{row}'].fill = savings_fill
    ws[f'E{row}'].font = Font(bold=True)
    apc_row = row
    row += 1
    
    ws[f'B{row}'] = 'SAVINGS'
    ws[f'B{row}'].font = Font(bold=True, size=12, color="006100")
    ws[f'E{row}'] = f'=E{trad_row}-E{apc_row}'
    ws[f'E{row}'].number_format = '$#,##0'
    ws[f'E{row}'].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    ws[f'E{row}'].font = Font(bold=True, size=12, color="006100")
    ws[f'F{row}'] = f'{((traditional_total - apc_total) / traditional_total * 100):.1f}% reduction'
    ws[f'F{row}'].font = Font(bold=True, color="006100")
    row += 2
    
    # Continue with GC costs using same percentages
    ws[f'B{row}'] = 'General Conditions'
    ws[f'C{row}'] = '=GC_GC'
    ws[f'C{row}'].number_format = '0.0%'
    ws[f'E{row}'] = f'=E{subtotal_row}*GC_GC'
    ws[f'E{row}'].number_format = '$#,##0'
    gc_row = row
    row += 1
    
    ws[f'B{row}'] = 'Overhead & Profit'
    ws[f'C{row}'] = '=GC_OHP'
    ws[f'C{row}'].number_format = '0.0%'
    ws[f'E{row}'] = f'=(E{subtotal_row}+E{gc_row})*GC_OHP'
    ws[f'E{row}'].number_format = '$#,##0'
    ohp_row = row
    row += 2
    
    # Before contingency
    ws[f'B{row}'] = 'Subtotal before Contingency'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'E{row}'] = f'=E{subtotal_row}+E{gc_row}+E{ohp_row}'
    ws[f'E{row}'].number_format = '$#,##0'
    ws[f'E{row}'].fill = total_fill
    before_cont_row = row
    row += 2
    
    # Contingencies
    ws[f'B{row}'] = 'Design Contingency'
    ws[f'C{row}'] = '=Design_Cont'
    ws[f'C{row}'].number_format = '0.0%'
    ws[f'E{row}'] = f'=E{before_cont_row}*Design_Cont'
    ws[f'E{row}'].number_format = '$#,##0'
    design_cont_row = row
    row += 1
    
    ws[f'B{row}'] = 'Construction Contingency'
    ws[f'C{row}'] = '=Constr_Cont'
    ws[f'C{row}'].number_format = '0.0%'
    ws[f'E{row}'] = f'=E{before_cont_row}*Constr_Cont'
    ws[f'E{row}'].number_format = '$#,##0'
    constr_cont_row = row
    row += 1
    
    ws[f'B{row}'] = 'Owner Contingency'
    ws[f'C{row}'] = '=Owner_Cont'
    ws[f'C{row}'].number_format = '0.0%'
    ws[f'E{row}'] = f'=E{before_cont_row}*Owner_Cont'
    ws[f'E{row}'].number_format = '$#,##0'
    owner_cont_row = row
    row += 2
    
    # TOTAL HARD COSTS
    ws[f'B{row}'] = 'TOTAL HARD COSTS (APC Filigree)'
    ws[f'B{row}'].font = Font(bold=True, size=12)
    ws[f'E{row}'] = f'=E{before_cont_row}+E{design_cont_row}+E{constr_cont_row}+E{owner_cont_row}'
    ws[f'E{row}'].number_format = '$#,##0'
    ws[f'E{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    ws[f'E{row}'].font = Font(bold=True, size=12)
    total_apc_row = row
    row += 1
    
    # Cost per GSF
    ws[f'B{row}'] = 'Cost per GSF'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'E{row}'] = f'=E{total_apc_row}/{TOTAL_GSF}'
    ws[f'E{row}'].number_format = '$#,##0.00'
    ws[f'E{row}'].font = Font(bold=True)
    row += 2
    
    # Comparison
    ws[f'B{row}'] = 'Traditional System Total'
    ws[f'E{row}'] = '=TotalHardCosts'
    ws[f'E{row}'].number_format = '$#,##0'
    trad_total_row = row
    row += 1
    
    ws[f'B{row}'] = 'APC Filigree Total'
    ws[f'E{row}'] = f'=E{total_apc_row}'
    ws[f'E{row}'].number_format = '$#,##0'
    ws[f'E{row}'].fill = savings_fill
    apc_total_row = row
    row += 1
    
    ws[f'B{row}'] = 'TOTAL PROJECT SAVINGS'
    ws[f'B{row}'].font = Font(bold=True, size=14, color="006100")
    ws[f'E{row}'] = f'=E{trad_total_row}-E{apc_total_row}'
    ws[f'E{row}'].number_format = '$#,##0'
    ws[f'E{row}'].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    ws[f'E{row}'].font = Font(bold=True, size=14, color="006100")
    ws[f'F{row}'] = f'=(E{trad_total_row}-E{apc_total_row})/E{trad_total_row}'
    ws[f'F{row}'].number_format = '0.0%'
    ws[f'F{row}'].font = Font(bold=True, size=12, color="006100")
    
    # Define named range
    wb.defined_names['TotalHardCosts_APC'] = DefinedName('TotalHardCosts_APC', attr_text=f"'Hard Costs - APC Filigree'!$E${total_apc_row}")

def create_apc_rfq_detail(wb):
    """Detailed RFQ for APC Filigree system"""
    ws = wb.create_sheet("RFQ - APC Filigree Detail")
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    ws['A1'] = 'REQUEST FOR QUOTE - APC FILIGREE PRECAST SYSTEM'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')
    
    ws['A2'] = 'NEVO Tower - 1580 79th St Causeway, North Bay Village, FL'
    ws.merge_cells('A2:H2')
    
    ws['A3'] = f'Total Project: {TOTAL_GSF:,} GSF | 7 Stories + Rooftop | 50 Residential Units'
    ws.merge_cells('A3:H3')
    
    ws['A4'] = 'Aerial Precast Concrete Composite Deck System'
    ws['A4'].font = Font(bold=True, italic=True)
    ws.merge_cells('A4:H4')
    
    row = 6
    
    # PROJECT OVERVIEW
    ws[f'A{row}'] = 'PROJECT OVERVIEW'
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:H{row}')
    row += 2
    
    overview = [
        ('Building Type:', 'Mixed-use residential tower'),
        ('Stories:', '7 stories + rooftop'),
        ('Total Area:', f'{TOTAL_GSF:,} GSF'),
        ('Structural System:', 'APC Filigree Composite Precast Deck'),
        ('Construction Duration:', '18 months (vs 24 traditional)'),
        ('Labor Reduction:', '30-40% vs traditional PT'),
    ]
    
    for label, value in overview:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        ws.merge_cells(f'B{row}:D{row}')
        row += 1
    
    row += 1
    
    # APC SYSTEM BENEFITS
    ws[f'A{row}'] = 'APC FILIGREE SYSTEM BENEFITS'
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:H{row}')
    row += 2
    
    benefits = [
        '✓ Semi-precast system reduces on-site labor by 30-40%',
        '✓ Eliminates traditional formwork and shoring',
        '✓ Faster construction schedule (25% time savings)',
        '✓ Improved quality control (factory fabrication)',
        '✓ Reduced concrete waste and material costs',
        '✓ Integrated structural steel support system',
        '✓ Fire-rated and code-compliant',
        '✓ Suitable for high-rise construction',
    ]
    
    for benefit in benefits:
        ws[f'A{row}'] = benefit
        ws[f'A{row}'].font = Font(color="006100")
        ws.merge_cells(f'A{row}:H{row}')
        row += 1
    
    row += 1
    
    # SCOPE OF WORK
    ws[f'A{row}'] = 'SCOPE OF WORK - APC FILIGREE SYSTEM'
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    headers = ['Item', 'Description', 'Quantity', 'Unit', 'Unit Price', 'Total', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    row += 1
    
    # Detailed scope items
    scope_items = [
        ('1', 'APC Filigree Deck Panels', TOTAL_GSF, 'SF', 28.50, '', 'Factory fabricated'),
        ('2', 'Structural Steel Integration', TOTAL_GSF, 'SF', 3.08, '', 'Embedded supports'),
        ('3', 'Topping Concrete', TOTAL_GSF * 0.25, 'CY', 200, '', '3" topping slab'),
        ('4', 'Panel Installation Labor', TOTAL_GSF, 'SF', 2.50, '', 'Crane & crew'),
        ('5', 'Shoring (minimal)', TOTAL_GSF * 0.2, 'SF', 1.50, '', 'Temporary only'),
        ('6', 'Connections & Anchors', 350, 'EA', 250, '', 'Per floor connections'),
        ('7', 'Engineering & Shop Drawings', 1, 'LS', 45000, '', 'Structural PE'),
        ('8', 'Transportation & Delivery', TOTAL_GSF / 1000, 'Loads', 2500, '', 'Per truck'),
        ('9', 'Crane Time', 90, 'Days', 1200, '', 'Tower crane rental'),
        ('10', 'Quality Control & Testing', 1, 'LS', 25000, '', 'Third party'),
    ]
    
    scope_start = row
    for item, desc, qty, unit, unit_price, total, notes in scope_items:
        ws[f'A{row}'] = item
        ws[f'B{row}'] = desc
        ws[f'C{row}'] = qty
        ws[f'C{row}'].number_format = '#,##0.00'
        ws[f'D{row}'] = unit
        ws[f'E{row}'] = unit_price
        ws[f'E{row}'].number_format = '$#,##0.00'
        ws[f'F{row}'] = f'=C{row}*E{row}'
        ws[f'F{row}'].number_format = '$#,##0'
        ws[f'G{row}'] = notes
        row += 1
    
    # Subtotal
    ws[f'B{row}'] = 'SUBTOTAL'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'F{row}'] = f'=SUM(F{scope_start}:F{row-1})'
    ws[f'F{row}'].number_format = '$#,##0'
    ws[f'F{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    ws[f'F{row}'].font = Font(bold=True)
    
    row += 2
    
    # SCHEDULE
    ws[f'A{row}'] = 'CONSTRUCTION SCHEDULE'
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:H{row}')
    row += 2
    
    schedule = [
        ('Phase 1', 'Shop Drawings & Fabrication', '2 months', 'Before on-site work'),
        ('Phase 2', 'Foundation & Columns', '3 months', 'Concurrent with fabrication'),
        ('Phase 3', 'Deck Installation (7 floors)', '9 months', '1.3 months per floor'),
        ('Phase 4', 'MEP & Finishes', '4 months', '25% overlap with structure'),
        ('Total Duration', 'Ground to CO', '18 months', '6 months faster than traditional'),
    ]
    
    for phase, desc, duration, notes in schedule:
        ws[f'A{row}'] = phase
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = desc
        ws.merge_cells(f'B{row}:D{row}')
        ws[f'E{row}'] = duration
        ws[f'E{row}'].font = Font(bold=True)
        ws[f'F{row}'] = notes
        ws.merge_cells(f'F{row}:H{row}')
        row += 1
    
    row += 1
    
    # SUBMITTAL REQUIREMENTS
    ws[f'A{row}'] = 'SUBMITTAL REQUIREMENTS'
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:H{row}')
    row += 2
    
    submittals = [
        '1. Complete project pricing breakdown',
        '2. Shop drawings and engineering calculations',
        '3. Product data sheets and specifications',
        '4. Quality control procedures',
        '5. Installation schedule and sequence',
        '6. Crane and equipment requirements',
        '7. Testing and inspection protocols',
        '8. Warranty information (minimum 2 years)',
        '9. References from similar high-rise projects',
        '10. Insurance certificates',
    ]
    
    for submittal in submittals:
        ws[f'A{row}'] = submittal
        ws.merge_cells(f'A{row}:H{row}')
        row += 1

if __name__ == "__main__":
    create_apc_workbook()
