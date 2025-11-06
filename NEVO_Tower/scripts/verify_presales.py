import openpyxl

wb = openpyxl.load_workbook('outputs/NEVO_Interactive_Budget_V3.xlsx')
ws = wb['24-Month Cash Flow']

print("\n" + "="*80)
print("PRE-SALES PAYMENT VERIFICATION")
print("="*80)

# Check headers
print(f"\nHeaders:")
print(f"  F: {ws['F4'].value}")
print(f"  G: {ws['G4'].value}")

# Check key months
print(f"\nPayment Schedule Formulas:")
for month in [1, 5, 15, 24]:
    row = 4 + month
    units = ws[f'E{row}'].value
    formula = ws[f'F{row}'].value
    usable = ws[f'G{row}'].value
    print(f"\nMonth {month}:")
    print(f"  Units Sold: {units}")
    print(f"  Payment Formula: {formula}")
    print(f"  Usable Formula: {usable}")

# Check verification row
print(f"\n" + "-"*80)
print("VERIFICATION ROW:")
for row in range(30, 35):
    if ws[f'A{row}'].value == 'VERIFICATION:':
        print(f"  {ws[f'B{row}'].value} {ws[f'C{row}'].value}")
        print(f"  {ws[f'D{row}'].value} {ws[f'E{row}'].value}")
        break

print("\n" + "="*80)
print("VERIFICATION COMPLETE")
print("="*80)
