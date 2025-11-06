#!/usr/bin/env python3
"""
Remove 'Copy of Measurements 50' and 'BUDGET' sheets from NEVO_Interactive_Budget_V3.xlsx
"""

import openpyxl

def main():
    # Load workbook
    target_path = r'C:\Users\17274\ME\2829-Niagara-Street\NEVO_Tower\outputs\NEVO_Interactive_Budget_V3.xlsx'
    wb = openpyxl.load_workbook(target_path)
    
    print(f"Sheets before: {wb.sheetnames}")
    
    # Remove unwanted sheets
    sheets_to_remove = ['Copy of Measurements 50', 'BUDGET']
    
    for sheet_name in sheets_to_remove:
        if sheet_name in wb.sheetnames:
            print(f"Removing: {sheet_name}")
            del wb[sheet_name]
    
    # Save
    wb.save(target_path)
    print(f"\n✓ Saved updated workbook")
    print(f"Sheets after: {wb.sheetnames}")

if __name__ == '__main__':
    main()
