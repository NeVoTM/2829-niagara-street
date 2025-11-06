"""
Comprehensive analysis of NEVO workbook to find discrepancies
"""
import openpyxl

def analyze_workbook():
    wb = openpyxl.load_workbook('outputs/NEVO_Interactive_Budget_V3.xlsx', data_only=True)
    
    print("=" * 80)
    print("NEVO TOWER WORKBOOK ANALYSIS")
    print("=" * 80)
    
    # SUMMARY SHEET
    print("\n" + "=" * 80)
    print("SUMMARY SHEET")
    print("=" * 80)
    ws = wb['Summary']
    sc_cash = ws['B5'].value or 0
    hard_costs = ws['B6'].value or 0
    soft_costs = ws['B7'].value or 0
    total_cost = ws['B8'].value or 0
    revenue = ws['B11'].value or 0
    profit = ws['B14'].value or 0
    
    print(f"SC Cash Payment:       ${sc_cash:>15,.2f}")
    print(f"Hard Costs:            ${hard_costs:>15,.2f}")
    print(f"Soft Costs:            ${soft_costs:>15,.2f}")
    print(f"TOTAL PROJECT COST:    ${total_cost:>15,.2f}")
    print(f"\nTotal Revenue:         ${revenue:>15,.2f}")
    print(f"Gross Profit:          ${profit:>15,.2f}")
    
    calc_total = sc_cash + hard_costs + soft_costs
    print(f"\nCalculated Total:      ${calc_total:>15,.2f}")
    print(f"Difference:            ${total_cost - calc_total:>15,.2f}")
    
    # ASSUMPTIONS SHEET
    print("\n" + "=" * 80)
    print("ASSUMPTIONS SHEET - LAND PARTNER")
    print("=" * 80)
    ws = wb['Assumptions']
    cash1 = ws['B12'].value or 0
    cash2 = ws['B13'].value or 0
    print(f"SC Key Money (Mo 0):   ${cash1:>15,.2f}")
    print(f"SC Payment (Mo 6):     ${cash2:>15,.2f}")
    print(f"Total SC Cash:         ${cash1 + cash2:>15,.2f}")
    
    # HARD COSTS SHEET
    print("\n" + "=" * 80)
    print("HARD COSTS SHEET")
    print("=" * 80)
    ws = wb['Hard Costs']
    # Find total row
    for i in range(30, 50):
        if ws[f'B{i}'].value and 'TOTAL HARD COSTS' in str(ws[f'B{i}'].value):
            total_hard_row = i
            break
    
    total_hard = ws[f'E{total_hard_row}'].value or 0
    print(f"TOTAL HARD COSTS:      ${total_hard:>15,.2f}")
    
    # SOFT COSTS SHEET
    print("\n" + "=" * 80)
    print("SOFT COSTS SHEET - DETAILED")
    print("=" * 80)
    ws = wb['Soft Costs']
    
    # Find subtotals
    inkind_total = None
    cash_total = None
    financing_total = None
    subtotal_before_financing = None
    grand_total = None
    
    for i in range(5, 35):
        cell_val = ws[f'A{i}'].value
        if cell_val:
            if 'SUBTOTAL IN-KIND' in str(cell_val):
                inkind_total = ws[f'B{i}'].value or 0
                print(f"SUBTOTAL IN-KIND:      ${inkind_total:>15,.2f}")
            elif 'SUBTOTAL CASH' in str(cell_val):
                cash_total = ws[f'B{i}'].value or 0
                print(f"SUBTOTAL CASH:         ${cash_total:>15,.2f}")
            elif 'Bridge Loan Interest' in str(cell_val):
                financing_total = ws[f'B{i}'].value or 0
                print(f"Bridge Loan Interest:  ${financing_total:>15,.2f}")
            elif 'SUBTOTAL SOFT COSTS (Before Financing)' in str(cell_val):
                subtotal_before_financing = ws[f'B{i}'].value or 0
                print(f"Subtotal Before Fin:   ${subtotal_before_financing:>15,.2f}")
            elif 'TOTAL SOFT COSTS (With Financing)' in str(cell_val):
                grand_total = ws[f'B{i}'].value or 0
                print(f"TOTAL SOFT (w/Fin):    ${grand_total:>15,.2f}")
    
    if inkind_total and cash_total:
        calc_subtotal = inkind_total + cash_total
        print(f"\nCalculated Subtotal:   ${calc_subtotal:>15,.2f}")
        if subtotal_before_financing:
            print(f"Difference:            ${subtotal_before_financing - calc_subtotal:>15,.2f}")
    
    # REVENUE SHEET
    print("\n" + "=" * 80)
    print("REVENUE SHEET")
    print("=" * 80)
    ws = wb['Revenue']
    for i in range(8, 15):
        if ws[f'B{i}'].value and 'TOTAL REVENUE' in str(ws[f'B{i}'].value):
            total_revenue = ws[f'F{i}'].value or 0
            units = ws[f'C{i}'].value or 0
            break
    
    print(f"Total Units:           {units:>20,.0f}")
    print(f"TOTAL REVENUE:         ${total_revenue:>15,.2f}")
    
    # Get synagogue
    synagogue_revenue = ws['F7'].value or 0
    print(f"Synagogue Revenue:     ${synagogue_revenue:>15,.2f}")
    print(f"50-Unit Revenue:       ${total_revenue - synagogue_revenue:>15,.2f}")
    
    # CASH FLOW SHEET - DETAILED
    print("\n" + "=" * 80)
    print("CASH FLOW SHEET - LINE BY LINE")
    print("=" * 80)
    ws = wb['24-Month Cash Flow']
    
    # Find totals row
    totals_row = None
    for i in range(28, 35):
        if ws[f'A{i}'].value and 'TOTALS' in str(ws[f'A{i}'].value):
            totals_row = i
            break
    
    if totals_row:
        print(f"\nTOTALS ROW: {totals_row}")
        hard_cf = ws[f'C{totals_row}'].value or 0
        cash_soft_cf = ws[f'D{totals_row}'].value or 0
        inkind_soft_cf = ws[f'E{totals_row}'].value or 0
        sc_cash_cf = ws[f'F{totals_row}'].value or 0
        units_sold = ws[f'G{totals_row}'].value or 0
        presales_total = ws[f'H{totals_row}'].value or 0
        usable_total = ws[f'I{totals_row}'].value or 0
        cash_in_total = ws[f'J{totals_row}'].value or 0
        net_flow_total = ws[f'K{totals_row}'].value or 0
        interest_total = ws[f'N{totals_row}'].value or 0
        
        print(f"\nHard Costs Total:      ${hard_cf:>15,.2f}")
        print(f"CASH Soft Total:       ${cash_soft_cf:>15,.2f}")
        print(f"IN-KIND Soft Total:    ${inkind_soft_cf:>15,.2f}")
        print(f"SC Cash Total:         ${sc_cash_cf:>15,.2f}")
        print(f"Interest Total:        ${interest_total:>15,.2f}")
        print(f"\nUnits Sold:            {units_sold:>20,.0f}")
        print(f"Pre-Sales Total:       ${presales_total:>15,.2f}")
        print(f"Usable Total:          ${usable_total:>15,.2f}")
        print(f"Cash IN Total:         ${cash_in_total:>15,.2f}")
        print(f"Net Flow Total:        ${net_flow_total:>15,.2f}")
        
        # Get Month 24 Final Cumulative
        final_cum = ws[f'O{totals_row-1}'].value or 0
        print(f"\nFinal Cumulative:      ${final_cum:>15,.2f}")
        
        # Find MAX BRIDGE and FINANCING rows
        for i in range(totals_row, totals_row + 5):
            if ws[f'A{i}'].value:
                if 'MAX BRIDGE' in str(ws[f'A{i}'].value):
                    max_bridge = ws[f'B{i}'].value or 0
                    print(f"\nMAX BRIDGE LOAN:       ${max_bridge:>15,.2f}")
                elif 'FINANCING COSTS' in str(ws[f'A{i}'].value):
                    financing_costs = ws[f'B{i}'].value or 0
                    print(f"FINANCING COSTS:       ${financing_costs:>15,.2f}")
    
    # RECONCILIATION
    print("\n" + "=" * 80)
    print("RECONCILIATION ANALYSIS")
    print("=" * 80)
    
    print("\n1. COST COMPONENTS:")
    print(f"   Hard Costs (Summary):     ${hard_costs:>15,.2f}")
    print(f"   Hard Costs (CF):          ${hard_cf:>15,.2f}")
    print(f"   Difference:               ${hard_costs - hard_cf:>15,.2f}")
    
    print(f"\n   Soft Costs (Summary):     ${soft_costs:>15,.2f}")
    if grand_total:
        print(f"   Soft Costs (Soft Sheet):  ${grand_total:>15,.2f}")
        print(f"   Difference:               ${soft_costs - grand_total:>15,.2f}")
    
    soft_from_cf = cash_soft_cf + inkind_soft_cf + interest_total
    print(f"   Soft from CF (C+I+Int):   ${soft_from_cf:>15,.2f}")
    print(f"   Difference:               ${soft_costs - soft_from_cf:>15,.2f}")
    
    print(f"\n   SC Cash (Summary):        ${sc_cash:>15,.2f}")
    print(f"   SC Cash (CF):             ${abs(sc_cash_cf):>15,.2f}")
    print(f"   Difference:               ${sc_cash - abs(sc_cash_cf):>15,.2f}")
    
    print("\n2. REVENUE CHECK:")
    print(f"   Revenue (Summary):        ${revenue:>15,.2f}")
    print(f"   Revenue (Revenue Sheet):  ${total_revenue:>15,.2f}")
    print(f"   Difference:               ${revenue - total_revenue:>15,.2f}")
    
    print(f"\n   Pre-Sales (CF):           ${presales_total:>15,.2f}")
    print(f"   Difference:               ${revenue - presales_total:>15,.2f}")
    
    print("\n3. PROFIT CHECK:")
    print(f"   Gross Profit (Summary):   ${profit:>15,.2f}")
    print(f"   Final Cumulative (CF):    ${final_cum:>15,.2f}")
    print(f"   Difference:               ${profit - final_cum:>15,.2f}")
    
    print("\n4. TOTAL PROJECT COST BUILD-UP:")
    total_calc_cf = hard_cf + cash_soft_cf + inkind_soft_cf + abs(sc_cash_cf) + interest_total
    print(f"   Hard + CASH Soft + IN-KIND + SC Cash + Interest")
    print(f"   ${hard_cf:,.2f} + ${cash_soft_cf:,.2f} + ${inkind_soft_cf:,.2f} + ${abs(sc_cash_cf):,.2f} + ${interest_total:,.2f}")
    print(f"   = ${total_calc_cf:>15,.2f}")
    print(f"   Summary Total:            ${total_cost:>15,.2f}")
    print(f"   DISCREPANCY:              ${total_cost - total_calc_cf:>15,.2f}")
    
    # Check if interest is double-counted
    print("\n5. INTEREST ANALYSIS:")
    print(f"   Interest in CF:           ${interest_total:>15,.2f}")
    print(f"   Financing in Soft Costs:  ${financing_total if financing_total else 0:>15,.2f}")
    print(f"   Are they equal?           {abs(interest_total - (financing_total or 0)) < 1}")
    
    # Final calculation check
    print("\n6. FINAL CUMULATIVE CALCULATION:")
    print(f"   Revenue:                  ${revenue:>15,.2f}")
    print(f"   - Total Costs:            ${total_cost:>15,.2f}")
    print(f"   = Expected Profit:        ${revenue - total_cost:>15,.2f}")
    print(f"   Actual Final Cum:         ${final_cum:>15,.2f}")
    print(f"   DIFFERENCE:               ${profit - final_cum:>15,.2f}")
    
    print("\n" + "=" * 80)
    print("END OF ANALYSIS")
    print("=" * 80)

if __name__ == "__main__":
    analyze_workbook()
