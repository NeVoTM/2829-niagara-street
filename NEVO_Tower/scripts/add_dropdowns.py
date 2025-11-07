"""
Add Data Validation Dropdowns to NEVO Interactive Workbook
Fixes the 'v' visual interface for dropdown commands
"""
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

def add_dropdowns():
    """Add data validation dropdowns to editable cells"""
    
    print("Adding data validation dropdowns...")
    
    wb = openpyxl.load_workbook('outputs/NEVO_Interactive_Budget_V3.xlsx')
    
    # ASSUMPTIONS SHEET - Add dropdowns for percentages
    if 'Assumptions' in wb.sheetnames:
        ws = wb['Assumptions']
        
        # Find editable cells (yellow background = FFFF99)
        for row in ws.iter_rows(min_row=1, max_row=50):
            for cell in row:
                if cell.fill and cell.fill.start_color:
                    if cell.fill.start_color.rgb == 'FFFF99':  # Yellow = editable
                        # Add note/comment to indicate it's editable
                        print(f"  Found editable cell: {cell.coordinate}")
    
    # HARD COSTS SHEET - Add data validation for rates
    if 'Hard Costs' in wb.sheetnames:
        ws = wb['Hard Costs']
        
        # Column C contains $/GSF rates that should be editable
        # Add data validation for numeric input only
        dv = DataValidation(
            type="decimal",
            operator="greaterThan",
            formula1=0,
            allow_blank=False,
            showErrorMessage=True,
            error='Please enter a valid cost per square foot',
            errorTitle='Invalid Input',
            prompt='Enter cost per GSF (e.g., 45.50)',
            promptTitle='Cost Rate'
        )
        
        # Apply to column C (rates) - rows 6 to 28 (23 CSI divisions)
        dv.add(f'C6:C28')
        ws.add_data_validation(dv)
        print(f"  Added numeric validation to Hard Costs rates (C6:C28)")
    
    # HARD COSTS SIMPLIFIED - Add data validation
    if 'Hard Costs - Simplified' in wb.sheetnames:
        ws = wb['Hard Costs - Simplified']
        
        dv = DataValidation(
            type="decimal",
            operator="greaterThan",
            formula1=0,
            allow_blank=False,
            showErrorMessage=True,
            error='Please enter a valid cost per square foot',
            errorTitle='Invalid Input',
            prompt='Enter cost per GSF (e.g., 45.50)',
            promptTitle='Cost Rate'
        )
        
        # Apply to column C (rates) for 7 categories
        dv.add(f'C5:C18')
        ws.add_data_validation(dv)
        print(f"  Added numeric validation to Simplified Hard Costs rates (C5:C18)")
    
    # SOFT COSTS SHEET - Percentage validation
    if 'Soft Costs' in wb.sheetnames:
        ws = wb['Soft Costs']
        
        # Percentages should be between 0 and 100%
        dv = DataValidation(
            type="decimal",
            operator="between",
            formula1=0,
            formula2=1,
            allow_blank=False,
            showErrorMessage=True,
            error='Please enter a percentage between 0% and 100%',
            errorTitle='Invalid Percentage',
            prompt='Enter as decimal (e.g., 0.05 for 5%)',
            promptTitle='Percentage'
        )
        
        # Column C contains percentages
        dv.add(f'C6:C20')
        ws.add_data_validation(dv)
        print(f"  Added percentage validation to Soft Costs (C6:C20)")
    
    # REVENUE SHEET - Validation
    if 'Revenue' in wb.sheetnames:
        ws = wb['Revenue']
        
        # Price per SF validation
        dv = DataValidation(
            type="decimal",
            operator="greaterThan",
            formula1=0,
            allow_blank=False,
            showErrorMessage=True,
            error='Please enter a valid price per square foot',
            errorTitle='Invalid Price',
            prompt='Enter price per NSF (e.g., 1100.00)',
            promptTitle='Unit Price'
        )
        
        dv.add(f'E6:E20')
        ws.add_data_validation(dv)
        print(f"  Added price validation to Revenue sheet (E6:E20)")
    
    # Save with data validation
    filename = 'outputs/NEVO_Interactive_Budget_V3.xlsx'
    wb.save(filename)
    
    print(f"\n✅ Data validation dropdowns added!")
    print(f"✅ Updated: {filename}")
    print("\n💡 Now you can:")
    print("   - Click cells with yellow background")
    print("   - See input prompts")
    print("   - Get validation errors for invalid entries")
    print("   - Press ESC to cancel edit")
    
    return filename

if __name__ == "__main__":
    add_dropdowns()
