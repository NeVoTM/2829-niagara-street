"""
Remove unwanted sheets from NEVO workbook
"""
import openpyxl

def remove_sheets():
    wb_path = 'outputs/NEVO_Interactive_Budget_V3.xlsx'
    
    print(f"Opening workbook: {wb_path}")
    wb = openpyxl.load_workbook(wb_path)
    
    sheets_to_remove = ['Copy of Measurements 50', 'BUDGET']
    
    print(f"\nSheets before cleanup: {wb.sheetnames}")
    
    for sheet_name in sheets_to_remove:
        if sheet_name in wb.sheetnames:
            print(f"  Removing: {sheet_name}")
            del wb[sheet_name]
        else:
            print(f"  Not found: {sheet_name}")
    
    print(f"\nSheets after cleanup: {wb.sheetnames}")
    
    wb.save(wb_path)
    print(f"\n✓ Saved cleaned workbook to: {wb_path}")

if __name__ == "__main__":
    remove_sheets()
