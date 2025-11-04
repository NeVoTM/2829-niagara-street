"""
Create realistic budget for 11-12 floor tower at 1580 79th Street
Target: Sell at $1,000/SF with reduced construction costs
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Configuration for shorter tower
TOWER_CONFIG = {
    'floors': 12,  # 11-12 floors (adjustable)
    'ground_floor': True,  # Lobby/amenity level
    'residential_floors': 11,
    'units_per_floor': 4,  # Reduced from larger tower
    'parking_ratio': 1.25,  # spaces per unit (reduced from typical 1.5-2.0)
}

# Unit mix - smaller, more efficient tower
UNIT_MIX = {
    '1BR': {'count': 16, 'avg_sf': 750, 'price_psf': 1050},  # Premium for smaller units
    '2BR': {'count': 20, 'avg_sf': 1100, 'price_psf': 1000},
    '3BR': {'count': 8, 'avg_sf': 1500, 'price_psf': 950},
}

# Construction costs for mid-rise (11-12 floors) - significantly less than high-rise
CONSTRUCTION_COSTS = {
    'hard_costs_psf': 325,  # Lower than high-rise ($400-500+), higher than low-rise ($250-300)
    'site_work': 850000,
    'parking_per_space': 35000,  # Surface or podium parking (not underground)
    'amenities': 750000,  # Modest: lobby, gym, roof deck
    'contingency_pct': 0.08,  # 8% contingency
}

# Soft costs
SOFT_COSTS = {
    'architecture_engineering_pct': 0.06,
    'permits_fees_pct': 0.03,
    'legal_accounting': 250000,
    'marketing_sales_pct': 0.03,
    'insurance_taxes': 180000,
    'financing_pct': 0.025,  # 2.5% of total project cost
    'developer_fee_pct': 0.05,  # 5% of hard costs
}

# Land/acquisition (adjust based on actual)
LAND_COST = 3500000  # Placeholder - adjust to actual

def calculate_project_budget():
    """Calculate complete project budget"""
    
    # Calculate total units and SF
    total_units = sum(unit['count'] for unit in UNIT_MIX.values())
    total_residential_sf = sum(unit['count'] * unit['avg_sf'] for unit in UNIT_MIX.values())
    avg_unit_size = total_residential_sf / total_units
    
    # Common areas (lobby, corridors, mechanical) - typically 15-20% of residential
    common_area_sf = int(total_residential_sf * 0.18)
    total_building_sf = total_residential_sf + common_area_sf
    
    # Parking calculation
    parking_spaces = int(total_units * TOWER_CONFIG['parking_ratio'])
    
    print(f"\n{'='*60}")
    print(f"1580 79TH STREET - {TOWER_CONFIG['floors']}-FLOOR TOWER BUDGET")
    print(f"{'='*60}\n")
    
    print(f"PROJECT SCOPE:")
    print(f"  Total Floors: {TOWER_CONFIG['floors']}")
    print(f"  Residential Floors: {TOWER_CONFIG['residential_floors']}")
    print(f"  Total Units: {total_units}")
    print(f"  Average Unit Size: {avg_unit_size:,.0f} SF")
    print(f"  Total Residential SF: {total_residential_sf:,.0f}")
    print(f"  Common Area SF: {common_area_sf:,.0f}")
    print(f"  Total Building SF: {total_building_sf:,.0f}")
    print(f"  Parking Spaces: {parking_spaces}")
    
    print(f"\n{'='*60}")
    print(f"UNIT MIX & REVENUE:")
    print(f"{'='*60}\n")
    
    total_revenue = 0
    for unit_type, data in UNIT_MIX.items():
        unit_revenue = data['count'] * data['avg_sf'] * data['price_psf']
        total_revenue += unit_revenue
        print(f"  {unit_type}: {data['count']} units @ {data['avg_sf']:,} SF × ${data['price_psf']}/SF")
        print(f"       Revenue: ${unit_revenue:,.0f} (${unit_revenue/data['count']:,.0f}/unit)")
    
    avg_price_psf = total_revenue / total_residential_sf
    print(f"\n  TOTAL GROSS REVENUE: ${total_revenue:,.0f}")
    print(f"  Average Price/SF: ${avg_price_psf:,.0f}")
    
    # HARD COSTS
    print(f"\n{'='*60}")
    print(f"HARD COSTS:")
    print(f"{'='*60}\n")
    
    base_construction = total_building_sf * CONSTRUCTION_COSTS['hard_costs_psf']
    site_work = CONSTRUCTION_COSTS['site_work']
    parking_cost = parking_spaces * CONSTRUCTION_COSTS['parking_per_space']
    amenities = CONSTRUCTION_COSTS['amenities']
    
    subtotal_hard = base_construction + site_work + parking_cost + amenities
    contingency = subtotal_hard * CONSTRUCTION_COSTS['contingency_pct']
    total_hard_costs = subtotal_hard + contingency
    
    print(f"  Base Construction: {total_building_sf:,} SF × ${CONSTRUCTION_COSTS['hard_costs_psf']}/SF")
    print(f"    = ${base_construction:,.0f}")
    print(f"  Site Work: ${site_work:,.0f}")
    print(f"  Parking: {parking_spaces} spaces × ${CONSTRUCTION_COSTS['parking_per_space']:,}")
    print(f"    = ${parking_cost:,.0f}")
    print(f"  Amenities: ${amenities:,.0f}")
    print(f"  Contingency ({CONSTRUCTION_COSTS['contingency_pct']*100}%): ${contingency:,.0f}")
    print(f"\n  TOTAL HARD COSTS: ${total_hard_costs:,.0f}")
    print(f"  Hard Cost/SF: ${total_hard_costs/total_building_sf:,.0f}")
    
    # SOFT COSTS
    print(f"\n{'='*60}")
    print(f"SOFT COSTS:")
    print(f"{'='*60}\n")
    
    arch_eng = total_hard_costs * SOFT_COSTS['architecture_engineering_pct']
    permits = total_hard_costs * SOFT_COSTS['permits_fees_pct']
    legal = SOFT_COSTS['legal_accounting']
    marketing = total_revenue * SOFT_COSTS['marketing_sales_pct']
    insurance = SOFT_COSTS['insurance_taxes']
    
    # Financing based on total project cost (iterative)
    project_before_financing = LAND_COST + total_hard_costs + arch_eng + permits + legal + marketing + insurance
    financing = project_before_financing * SOFT_COSTS['financing_pct']
    
    developer_fee = total_hard_costs * SOFT_COSTS['developer_fee_pct']
    
    total_soft_costs = arch_eng + permits + legal + marketing + insurance + financing + developer_fee
    
    print(f"  Architecture/Engineering ({SOFT_COSTS['architecture_engineering_pct']*100}%): ${arch_eng:,.0f}")
    print(f"  Permits & Fees ({SOFT_COSTS['permits_fees_pct']*100}%): ${permits:,.0f}")
    print(f"  Legal & Accounting: ${legal:,.0f}")
    print(f"  Marketing & Sales ({SOFT_COSTS['marketing_sales_pct']*100}%): ${marketing:,.0f}")
    print(f"  Insurance & Taxes: ${insurance:,.0f}")
    print(f"  Financing Costs ({SOFT_COSTS['financing_pct']*100}%): ${financing:,.0f}")
    print(f"  Developer Fee ({SOFT_COSTS['developer_fee_pct']*100}%): ${developer_fee:,.0f}")
    print(f"\n  TOTAL SOFT COSTS: ${total_soft_costs:,.0f}")
    
    # TOTAL PROJECT COST
    print(f"\n{'='*60}")
    print(f"TOTAL PROJECT COST:")
    print(f"{'='*60}\n")
    
    total_project_cost = LAND_COST + total_hard_costs + total_soft_costs
    cost_per_sf = total_project_cost / total_residential_sf
    cost_per_unit = total_project_cost / total_units
    
    print(f"  Land/Acquisition: ${LAND_COST:,.0f}")
    print(f"  Hard Costs: ${total_hard_costs:,.0f}")
    print(f"  Soft Costs: ${total_soft_costs:,.0f}")
    print(f"\n  TOTAL PROJECT COST: ${total_project_cost:,.0f}")
    print(f"  Cost per SF: ${cost_per_sf:,.0f}")
    print(f"  Cost per Unit: ${cost_per_unit:,.0f}")
    
    # PRO FORMA
    print(f"\n{'='*60}")
    print(f"PRO FORMA SUMMARY:")
    print(f"{'='*60}\n")
    
    gross_profit = total_revenue - total_project_cost
    profit_margin = (gross_profit / total_revenue) * 100
    roi = (gross_profit / total_project_cost) * 100
    
    print(f"  Total Revenue: ${total_revenue:,.0f}")
    print(f"  Total Cost: ${total_project_cost:,.0f}")
    print(f"  Gross Profit: ${gross_profit:,.0f}")
    print(f"  Profit Margin: {profit_margin:.1f}%")
    print(f"  Return on Cost: {roi:.1f}%")
    print(f"\n  Profit per Unit: ${gross_profit/total_units:,.0f}")
    print(f"  Profit per SF: ${gross_profit/total_residential_sf:,.0f}")
    
    # KEY METRICS
    print(f"\n{'='*60}")
    print(f"KEY METRICS:")
    print(f"{'='*60}\n")
    
    print(f"  Revenue/SF: ${avg_price_psf:,.0f}")
    print(f"  Cost/SF: ${cost_per_sf:,.0f}")
    print(f"  Spread: ${avg_price_psf - cost_per_sf:,.0f}/SF")
    print(f"  Cost as % of Revenue: {(total_project_cost/total_revenue)*100:.1f}%")
    
    # Create Excel workbook
    create_excel_budget(
        total_units, total_residential_sf, total_building_sf, parking_spaces,
        total_revenue, total_hard_costs, total_soft_costs, total_project_cost,
        gross_profit, profit_margin, roi
    )
    
    return {
        'total_revenue': total_revenue,
        'total_cost': total_project_cost,
        'profit': gross_profit,
        'margin': profit_margin
    }

def create_excel_budget(units, res_sf, bldg_sf, parking, revenue, hard, soft, total, profit, margin, roi):
    """Create detailed Excel budget"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget Summary"
    
    # Styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    section_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    section_font = Font(bold=True, size=11)
    
    row = 1
    
    # Title
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = f"1580 79TH STREET - {TOWER_CONFIG['floors']}-FLOOR TOWER"
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center')
    row += 1
    
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = f"Budget Date: {datetime.now().strftime('%B %d, %Y')}"
    cell.alignment = Alignment(horizontal='center')
    row += 2
    
    # Project Scope
    ws[f'A{row}'] = "PROJECT SCOPE"
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    row += 1
    
    scope_data = [
        ("Total Units", units),
        ("Residential SF", f"{res_sf:,}"),
        ("Total Building SF", f"{bldg_sf:,}"),
        ("Parking Spaces", parking),
        ("Floors", TOWER_CONFIG['floors']),
    ]
    
    for label, value in scope_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1
    
    row += 1
    
    # Revenue Summary
    ws[f'A{row}'] = "REVENUE"
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    for unit_type, data in UNIT_MIX.items():
        ws[f'A{row}'] = f"{unit_type} Units"
        ws[f'B{row}'] = data['count']
        ws[f'C{row}'] = f"{data['avg_sf']:,} SF"
        ws[f'D{row}'] = f"${data['count'] * data['avg_sf'] * data['price_psf']:,.0f}"
        row += 1
    
    ws[f'A{row}'] = "TOTAL REVENUE"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'D{row}'] = f"${revenue:,.0f}"
    ws[f'D{row}'].font = Font(bold=True)
    row += 2
    
    # Costs
    ws[f'A{row}'] = "PROJECT COSTS"
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    costs = [
        ("Land/Acquisition", LAND_COST),
        ("Hard Costs", hard),
        ("Soft Costs", soft),
    ]
    
    for label, amount in costs:
        ws[f'A{row}'] = label
        ws[f'D{row}'] = f"${amount:,.0f}"
        row += 1
    
    ws[f'A{row}'] = "TOTAL COST"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'D{row}'] = f"${total:,.0f}"
    ws[f'D{row}'].font = Font(bold=True)
    row += 2
    
    # Summary
    ws[f'A{row}'] = "PRO FORMA"
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    ws[f'A{row}'] = "Gross Profit"
    ws[f'D{row}'] = f"${profit:,.0f}"
    row += 1
    
    ws[f'A{row}'] = "Profit Margin"
    ws[f'D{row}'] = f"{margin:.1f}%"
    row += 1
    
    ws[f'A{row}'] = "Return on Cost"
    ws[f'D{row}'] = f"{roi:.1f}%"
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    
    # Save
    filename = f"1580_Tower_{TOWER_CONFIG['floors']}Floor_Budget.xlsx"
    wb.save(filename)
    print(f"\n{'='*60}")
    print(f"Excel budget saved: {filename}")
    print(f"{'='*60}\n")

if __name__ == "__main__":
    calculate_project_budget()
